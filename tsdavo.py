import requests
import json
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from googletrans import Translator
import re
import string
from time import sleep
from datetime import date
from httpcore._exceptions import ReadTimeout, ConnectTimeout

base = 'https://e-resource.tsdavo.gov.ua'
cache_dir = './cache'

def clean(msg):
    # remove all but alphanumeric, hyphen and space
    allow = string.ascii_letters + string.digits + ' ' + '-'
    return re.sub('[^%s]' % allow, '', msg)

def is_english(s):
    try:
        s.encode(encoding='utf-8').decode('ascii')
    except UnicodeDecodeError:
        return False
    else:
        return True

translator = Translator()
def translation(text):
    if isinstance(text, (list, tuple)):
        return [translation(item) for item in text]
    print('translating: ', text)
    result = None
    wait_time = 1.
    for i in range(5):
        try:
            result = translator.translate(text, src='uk', dest='en')
            break
        except (requests.Timeout, ReadTimeout, ConnectTimeout) as err:
            print('translation timeout. retrying...')
        sleep(wait_time)
        wait_time *= 2
    assert result is not None 
    return result.text

def translate_field(items, field_name):
    batch = [item[field_name] for item in items]
    batch = translation(batch)
    for i, text in enumerate(batch):
        items[i][field_name] = text    

def load_fond(url):
    soup = BeautifulSoup(requests.get(url).text, 'lxml')
    result = []
    for tag in soup.find_all('div', attrs = {'class': 'row with-border-bottom thin-row'}):
        title = tag.find('div', attrs = {'class': 'left'})
        item = {'link': title.a["href"], 'title': title.text.strip()}
        result.append(item)
    # look for description
    historical_reference = 'Історична довідка'
    description = None
    for tag in soup.find_all('div', attrs = {'class': 'left'}):
        item = tag.text
        if item == historical_reference:
            tag = tag.find_next_sibling()
            description = translation(tag.text)
            break
    return result, description

def index_fonds(fonds):
    index = {}
    for item in fonds:
        index[item['number']] = item
    return index

def save_cached_object(obj, fname):
    fname = f'{cache_dir}/{fname}'
    with open(fname, 'w') as f:
        f.write(json.dumps(obj))

def load_cached_object(fname):
    fname = f'{cache_dir}/{fname}'
    with open(fname) as f:
        return json.loads(f.read())

def clean(msg):
    # remove all but alphanumeric, hyphen and space
    allow = string.ascii_letters + string.digits + ' ' + '-'
    return re.sub('[^%s]' % allow, '', msg)

class FondCollection:
    def __init__(self):
        self._base = 'https://e-resource.tsdavo.gov.ua'
        self._endpoint = '/api/v1/fonds/'
        self._sortfield = 'FondNumber'
        self._sortorder = 'asc'
        self._fonds = None
        self._fond_index = None
    
    def load_page(self, page=1, limit=20, translate=True):
        url = f'{self._base}{self._endpoint}?Limit={limit}&Page={page}&SortField={self._sortfield}&SortOrder={self._sortorder}'
        data = requests.get(url)
        result = json.loads(data.text)
        soup = BeautifulSoup(result['View'], 'lxml')
        items = []
        for tag in soup.find_all('td', attrs = {'class': 'name'}):
            date = tag.find_next_sibling()
            pages = date.find_next_sibling()
            number = tag.find_previous_sibling()
            title = tag.text.strip()
            item = { 'number': number.text,
                     'link': tag.a["href"], 
                     'title': title, 
                     'date': date.text,
                     'pages': pages.text}
            items.append(item)
        if translate:
            translate_field(items, 'title')
        return items
    
    def load(self, translate=True):
        try:
            result = load_cached_object('fonds.json')
        except:
            result = []
            limit = 1000
            page = 1
            while True:
                print('loading page', page)
                chunk = self.load_page(page=page, limit=limit, translate=translate)
                if not chunk: break
                print(f'  got {len(chunk)} items')
                result += chunk
                page += 1
            save_cached_object(result, 'fonds.json')
        result.sort(key=lambda x: int(x['number']))
        self._fonds = result
        self._fond_index = index_fonds(self._fonds)
        return result

    def lookup(self, fond_name):
        if not self._fond_index:
            self.load()
        return self._fond_index[fond_name]

collection = FondCollection()

class OpusInventory:
    def __init__(self, endpoint):
        self._base = 'https://e-resource.tsdavo.gov.ua'
        self._endpoint = '/api/v1' + endpoint
    
    def load_page(self, page=1, limit=20, translate=True):
        url = f'{self._base}{self._endpoint}?Limit={limit}&Page={page}'
        print('loading', url)
        data = requests.get(url)
        result = json.loads(data.text)
        soup = BeautifulSoup(result['View'], 'lxml')
        items = []
        for tag in soup.find_all('div', attrs = {'class': 'row with-border-bottom column'}):
            #print(tag)
            left = tag.find('div', attrs = {'class': 'left'})
            link = left.a['href']
            title = left.a.text.strip()
            date = left.find('span', attrs = {'class': 'date'}).text
            right = tag.find('div', attrs = {'class': 'right'})
            desc = right.text.strip()
            items.append({
                'link': link,
                'title': title,
                'date': date,
                'description': desc
            })
        if translate:
            translate_field(items, 'title')
            translate_field(items, 'description')
        return items
        
    def load_all(self, translate=True):
        result = []
        limit = 1000
        page = 1
        while True:
            print('loading page', page)
            chunk = self.load_page(page=page, limit=limit, translate=translate)
            if not chunk: break
            print(f'  got {len(chunk)} items')
            result += chunk
            page += 1
        return result

class Fond:
    def __init__(self, name=None):
        self._name = name
        self._url = None
        self._dates = None
        self._num_opi = None
        self._num_cases = None
        self._description = None
        self._opi = None

        if name is not None:
            self.load()

    Fields = {
        'Фонд':                 '_name',
        'Дати':                 '_dates',
        'Кількість описів':     '_num_opi',
        'Кількість справ':      '_num_cases',
        'Історична довідка':    '_description'
    }

    StoredFields = ['_name', '_url', '_dates', '_num_opi', '_num_cases', '_description', '_opi']

    def load(self):
        fond_info = collection.lookup(self._name)
        try:
            data = load_cached_object(f'fond{self._name}.json')
            for f in Fond.StoredFields:
                setattr(self, f, data[f])
            return
        except:
            pass

        url = base + fond_info['link']
        soup = BeautifulSoup(requests.get(url).text, 'lxml')
        self._url = url

        # load fond field values
        description = None
        for tag in soup.find_all('div', attrs = {'class': 'left'}):
            item = tag.text
            if item in Fond.Fields:
                tag = tag.find_next_sibling()
                value = tag.text
                if len(value) and not is_english(value):
                    value = translation(value)
                setattr(self, Fond.Fields[item], value)

        # load opus specifications
        result = []
        for tag in soup.find_all('div', attrs = {'class': 'row with-border-bottom thin-row'}):
            title = tag.find('div', attrs = {'class': 'left'})
            item = {'link': title.a["href"], 'title': title.text.strip()}
            result.append(item)
        self._opi = result

        data = {}
        for f in Fond.StoredFields:
            data[f] = getattr(self, f)
        save_cached_object(data, f'fond{self._name}.json')

def are_pages_linked(case_info):
    # check to see if the case has linked documents or not
    # the test is based on whether the button at the bottom of the page 
    # says "order" which means they are not linked, 
    # or "revision" which means they linked
    order = 'Замовити'
    revision = 'Перегляд'
    url = base + case_info['link']
    soup = BeautifulSoup(requests.get(url).text, 'lxml')
    for tag in soup.find_all('a', attrs = {'class': 'continue-link'}):
        item = tag.text.strip()
        if item == revision:
            return True
        if item == order:
            return False
    return None

class OpusTable:
    def __init__(self, fond_id, opus_index=1):
        self._fond_id = fond_id
        self._opus_index = opus_index
        self._opus_id = f'{fond_id}-{opus_index}'
        self._fond_description = None
        self._abstract = None
        self._cases = None
        self._linked_pages = None

    def load(self, base=base):
        try:
            self._cases = load_cached_object(f'opus{self._opus_id}.json')
            self._fond_description = load_cached_object(f'fond{self._fond_id}_description.json')
            self._abstract = load_cached_object(f'opus{self._opus_id}_abstract.json')
            return self
        except:
            pass
        fond_data = collection.lookup(self._fond_id)
        url = base + fond_data['link']
        print(f'loading {self._fond_id} from {url}')
        opi, self._fond_description = load_fond(url)
        url = opi[self._opus_index-1]['link']
        print(f'loading {self._opus_id} from {url}')
        inv = OpusInventory(url)
        self._cases = inv.load_all()
        self._abstract = self.load_abstract(base + url)
        save_cached_object(self._cases, f'opus{self._opus_id}.json')
        save_cached_object(self._fond_description, f'fond{self._fond_id}_description.json')
        save_cached_object(self._abstract, f'opus{self._opus_id}_abstract.json')
        return self

    def load_abstract(self, url):
        #print('looking for opus abstract:', url)
        soup = BeautifulSoup(requests.get(url).text, 'lxml')
        abstract = 'Анотація'
        for tag in soup.find_all('div', attrs = {'class': 'left bold'}):
            if tag.text == abstract:
                tag = tag.find_next_sibling()
                abstract = translation(tag.text) if len(tag.text) > 0 else ''
                #print('found abstract:', abstract)
                return abstract
        return None

    def scan_for_keywords(self, keywords):
        self.load()
        keywords = [clean(k) for k in keywords]
        pattern = '|'.join(keywords)
        expr = re.compile(pattern, re.IGNORECASE)
        opus_scan = False
        if self._fond_description and re.search(expr, self._fond_description):
            opus_scan = True
        elif self._abstract and re.search(expr, self._abstract):
            opus_scan = True
        case_scan = [re.search(expr, case['title']) is not None or re.search(expr, case['description']) is not None for case in self._cases]
        return opus_scan, case_scan

    def linked_pages(self):
        if self._linked_pages is None:
            try:
                self._linked_pages = load_cached_object(f'opus{self._opus_id}_linked_pages.json')
            except:
                self.load()
                print(f'checking {len(self._cases)} cases on the archive...')
                self._linked_pages = [are_pages_linked(case) for case in self._cases]
                for i, link in enumerate(self._linked_pages):
                    self._cases[i]['linked'] = link
                save_cached_object(self._cases, f'opus{self._opus_id}.json')
                save_cached_object(self._linked_pages, f'opus{self._opus_id}_linked_pages.json')
        return self._linked_pages

    def format_case(self, case, base=base):
        link = case['link']
        num = re.sub('Case ', '', case['title'])
        link = base + link
        desc = case['description']
        return (num, desc, link, case['date'])

    def add_row(self, row, case, base=base):
        num, desc, link, date = self.format_case(case, base)
        #print(num, title, link)
        row[0].value = num
        row[0].hyperlink = link
        row[0].style = 'Hyperlink'
        row[1].value = desc
        row[1].hyperlink = link
        row[1].style = 'Hyperlink'
        row[2].value = date
        linked = case['linked']
        row[4].value = "Linked" if linked is True else "Unlinked" if linked is False else ""

    def export(self, fname=None, case_filter=None):
        if not fname:
            fname = f'TSDAVO Opus {self._opus_id}.xlsx'
        self.linked_pages() # make sure linked info is loaded
        wb = load_workbook(filename = 'Opus Sample.xlsx')
        sheet = wb.active
        sheet.title = f"CDAVO {self._opus_id}"
        sheet["A1"].value = f"Archives / CDAVO / {self._fond_id} / {self._opus_index}"
        sheet["A2"].value = f"{self._opus_index} ..."
        sheet["D3"].value = "URL NEEDED!"
        sheet["D4"].value = date.today().strftime('%d %b %Y')
        #sheet.append(['Number', 'Description', 'Date'])
        cur_row = 9
        for i, item in enumerate(self._cases):
            if not case_filter or case_filter[i]:
                self.add_row(sheet[cur_row], item)
                cur_row += 1
        last_row = sheet.max_row
        #formula_row = sheet.max_row + 2
        sheet.append([])
        sheet.append([
            f"=rows(A9:a{last_row})",
            "totals",
            "",
            "",
            f"=counta(E9:E{last_row})",
            "",
            "",
            "",
            f"=counta(I9:I{last_row})",
            f"=counta(J9:J{last_row})",
            f"=sum(K9:K{last_row})",
            f"=counta(L9:L{last_row})",
            f"=counta(M9:M{last_row})",
            f"=sum(N9:N{last_row})",
            ])
        print(f'saving {sheet.max_row} rows to {fname}')
        wb.save(fname)


