
import re
import string
from copy import copy
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Color, Border
from utility import get_text, ARCHIVE_BASE


def child_url(child):
    link = child[0]["link"]
    return ARCHIVE_BASE + link if link is not None else None

def link_status(url):
    return "unlinked" if url is None or "redlink" in url else "linked"

EXPR_PATTERN = re.compile(r'{[^}]+}')

def check_string(s):
    if not s: return None
    match = re.findall(EXPR_PATTERN, s)
    if not match: return None
    return [m for m in match]

def check_cell(cell):
    return check_string(cell.value)

PARSE_PATTERN = re.compile(r'{(?P<expr>[a-zA-Z0-9_.]+)(\[(?P<index>[0-9]+)\])?(\:(?P<modifier>[a-zA-Z_]+))?}')

def parse_template_expr(expr):
    match = re.match(PARSE_PATTERN, expr)
    return match.groupdict() if match else None

def substitute(page, expr):
    try:
        fstring = '{__page__.' + expr['expr'] + '}'
        return string.Formatter().format(fstring, __page__=page)
    except:
        return None

def export_page(page, dest_file=None):
    template_file = f'templates/{page.kind}.xlsx'
    print(f'opening template file {template_file}...')
    wb = load_workbook(filename = template_file)
    sheet = wb.active

    # process title
    title = sheet.title
    check = check_string(title)
    if check:
        for expr in check:
            parsed = parse_template_expr(expr)
            sub = substitute(page, parsed)
            if sub:
                title = title.replace(expr, sub)
    print('sheet title:', title)
    sheet.title = title

    # make a list of all cells with template expressions
    edit_cell = {}
    edits = []
    for row in sheet.iter_rows():
        for cell in row:
            check = check_cell(cell)
            if check:
                parsed = [parse_template_expr(e) for e in check]
                for parse in parsed:
                    # check for formatting directives (before actually editing)
                    if parse['expr'] == 'edit':
                        # {edit} cells contain formatting for editing highlights
                        #print('edit cell found:', cell.coordinate, parse['modifier'])
                        edit_cell[parse['modifier']] = cell
                edits.append((cell, check, parsed))
    
    first_child_row = None
    last_child_row = None
    #rollup_font = Font(bold=True)
    #border = Border(outline=True)
    columns = []

    for edit in edits:
        cell, matches, parses = edit
        for match, parse in zip(matches, parses):
            #print('processing match:', match, parse)
            if parse['expr'] == 'col':
                # defer {col} expressions until after other substitutions are done
                # since this fills in the table and will overwrite the {rollup} expressions
                first_child_row = row = cell.row
                last_child_row = first_child_row + len(page.children) - 1
                columns.append((cell, parse))
            elif parse['expr'] == 'rollup':
                rollup_cell = sheet.cell(row=cell.row - 1 + len(page.children), column=cell.column)
                col = cell.column_letter
                rollup_cell.value = f'={parse["modifier"]}({col}{first_child_row}:{col}{last_child_row})'
                rollup_cell.style = cell.style
                rollup_cell.border = copy(cell.border)
                rollup_cell.alignment = copy(cell.alignment)
                rollup_cell.font = copy(cell.font)
                cell.value = ""
            elif parse['expr'] == 'edit':
                # {edit} cells contain formatting for editing highlights (caught above)
                pass
            else:
                # general case: replace template expression with substitution value
                sub = substitute(page, parse)
                if sub is not None:
                    cell.value = cell.value.replace(match, sub)
            if parse['modifier'] == 'linked':
                cell.hyperlink = page.url

    for cell, parse in columns:
        row = cell.row
        col = cell.column
        index = parse['index']
        for child in page.children:
            child_cell = sheet.cell(row=row, column=col)
            if child_cell.row != cell.row:
                # propagate border, style, font, and alignment to all rows
                child_cell.style = cell.style
                child_cell.border = copy(cell.border)
                child_cell.alignment = copy(cell.alignment)
                child_cell.font = copy(cell.font)
            if index is not None:
                item = child[int(index)]
                sub = get_text(item['text'])
                if 'edit' in item:
                    edit = item['edit']
                    if edit in edit_cell:
                        #print('edited:', item)
                        child_cell.fill = copy(edit_cell[edit].fill)
                child_cell.value = sub
            if parse['modifier'] == 'linked':
                child_cell.hyperlink = child_url(child)
            elif parse['modifier'] == 'link_status':
                child_cell.value = link_status(child_url(child))
            row += 1

    if dest_file:
        wb.save(dest_file)
    return wb
