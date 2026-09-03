import sys

# Remove the problematic jupyter virtual environment path if it gets auto-injected
sys.path = [p for p in sys.path if "venv-jupyter" not in p]
import re

import openpyxl
from file_location import (
    FileLocationFinder,
    get_unique_random_integers,
    remove_words_list,
)


class LocationPerformanceEvaluator:
    def __init__(self, file_path: str):
        """
        Reads an Excel file and returns the string values of cells C{N} and G{N}.
        """
        # Load the workbook
        workbook = openpyxl.load_workbook(file_path, data_only=True)

        # Select the active sheet
        self.sheet = workbook.active
        if not self.sheet:
            raise RuntimeError(f"failure reading the spreadsheet {file_path}")

        # Total rows minus 1 header row
        self.total_data_rows = self.sheet.max_row - 1
        self.file_location_finder = FileLocationFinder()
        self.num_evaluated_docs = 0
        self.num_docs_evaluated_correctly = 0
        self.num_docs_with_all_locations_found = 0
        self.total_num_cumulative_locs = 0
        self.total_num_cumulative_locs_in_jgdb = 0
        self.total_num_extracted_locs = 0
        self.total_num_coinciding_locs = 0


    def get_file_towns_from_cumulative_report(self, row_number: int):
        # Retrieve cells using standard Excel indexing coordinates (Column C "File" is 3, G "Town(s)" is 7)
        if not self.sheet:
            return None, None

        if row_number > self.total_data_rows or row_number < 2:
            raise ValueError(f"Invalid document number {row_number - 1}, "
                             f"it must be between 1 and {self.total_data_rows - 1}")

        cell_c_value = self.sheet.cell(row=row_number, column=3).value
        cell_g_value = self.sheet.cell(row=row_number, column=7).value

        # Cast to string safely, treating empty cells (None) as empty strings
        file = "" if cell_c_value is None else str(cell_c_value)
        towns = "" if cell_g_value is None else str(cell_g_value)

        words_to_delete = ["Koloniy", "Volost'", "Village", "Khutor", "a", "at", "the", "nr.", "near",
                           "monastery", "sugar", "beet", "plant", "big", "road", "Fabrica", "forest",
                           "station", "railway", "beh.bridge", "sloboda", "slobodka"]

        town_list = []
        if towns:
            # Splits by either (a comma followed by zero or more spaces) OR (a pipe symbol)
            # [,\s]+ matches commas and spaces, and " I " with a capital I
            town_list = re.split(r",\s*|\s+I\s+", towns)
        # interpret brackets as separators
        no_brackets_town_list = []
        for town in town_list:
            match = re.match(r"\s*(.*?)\s*\(\s*(.*?)\s*\)\s*", town)
            if match:
                no_brackets_town_list.extend([match.group(1), match.group(2)])
            else:
                no_brackets_town_list.append(town)

        town_list = [
            cleaned_town
            for town in no_brackets_town_list
            if (cleaned_town := remove_words_list(town, words_to_delete).strip())
        ]

        return file, town_list

    def get_doc_labels(self, file: str):
        """
        Splits the string at the first space character.
        Returns the part before and the part after the space.
        """
        # Clean leading/trailing spaces
        file = file.strip()

        split_parts = re.split(r"[ /]", file, maxsplit=1)
        if len(split_parts) < 2:
            raise ValueError(
                f"The provided string {file} does not contain any space or slash characters."
            )
        archive_name, fund_hyphen_etc = split_parts

        # Replaces all occurrences of '-' with '/' in the input string.
        fund_slash_etc = fund_hyphen_etc.replace("-", "/")
        if re.search(r'[\u0400-\u04FF]', archive_name):
            # Cyrillic archive name to Latin
            latin_archive_name = self.file_location_finder.find_archive_name_by_cyrillic_abbr(archive_name)
            if not latin_archive_name:
                raise ValueError(f"The Cyrillic abbreviation {archive_name} was not found.")
        else:
            latin_archive_name = archive_name
        suffixes = ['/', '-D/', '-R/', '-K/', '-P/', '-N/', '-A/']
        possible_labels = [latin_archive_name + suffix + fund_slash_etc for suffix in suffixes]
        return possible_labels

    def evaluate_on_one_doc(self, doc_id: int, cumulative_towns: list[str], file: str, 
                            skip_extraction:bool = False, debug_print:bool = False):
        # towns from the cumulative report to JG location IDs
        cumulative_locations = self.file_location_finder.match_places_to_location_ids(
            doc_id, cumulative_towns, debug_print)
        # Convert the frozen sets to dictionaries
        cumulative_locations = [dict(f_set) for f_set in cumulative_locations]
        # get location IDs
        cumulative_locations_ids = [loc["loc_id"] for loc in cumulative_locations]
        cumulative_locations_ids_set = set(cumulative_locations_ids)
        if 0 == len(cumulative_locations_ids_set):
            print(f"None of the towns {cumulative_towns} for document {file} is contained in the JGDB")
            return

        if skip_extraction:
            doc_location_ids = []
        else:
            doc_location_ids = self.file_location_finder.get_doc_location(doc_id,
                only_smallest_locations=True, debug_print=debug_print)

        doc_location_ids_set = set(doc_location_ids)
        if doc_location_ids_set == cumulative_locations_ids_set:
            self.num_docs_evaluated_correctly += 1
        if cumulative_locations_ids_set.issubset(doc_location_ids_set):
            self.num_docs_with_all_locations_found += 1

        if debug_print:
            if doc_location_ids_set == cumulative_locations_ids_set:
                msg = f"Record {file}, document ID {doc_id}, all locations coincide: "
            else:
                msg = f"Record {file}, document ID {doc_id}, {len(cumulative_locations_ids_set)} assumed locations: "
                for loc_id in cumulative_locations_ids_set:
                    location = self.file_location_finder.get_location_from_id(str(loc_id))
                    if location:
                        msg = f"{msg} {location["main_name"]}"
                msg = f"{msg}; identified {len(doc_location_ids_set)}:"
            for loc_id in doc_location_ids_set:
                location = self.file_location_finder.get_location_from_id(str(loc_id))
                if location:
                    msg = f"{msg} {location["main_name"]}"
            print(msg)

        intersection_set = doc_location_ids_set & cumulative_locations_ids_set
        
        self.num_evaluated_docs += 1
        self.total_num_cumulative_locs += len(set(cumulative_towns))
        self.total_num_cumulative_locs_in_jgdb += len(cumulative_locations_ids_set)
        self.total_num_extracted_locs += len(doc_location_ids_set)
        self.total_num_coinciding_locs += len(intersection_set)

    def unique_random_integers(self, max_num_docs: int) -> list[int]:
        rows = get_unique_random_integers(max_num_docs, self.total_data_rows)
        rows = [row + 1 for row in rows]  # the first row is the header
        return rows

    def evaluate_location_extraction(
        self,
        rows: list[int],
        skip_extraction: bool = False,
        debug_print: bool = False,
        batch_size: int = 20,
    ):
        max_num_docs = len(rows)
        rows = [row + 1 for row in rows]  # the first row is the header

        # Buffer: (row, file, town_list, doc_id) per document
        pending: list[tuple[int, str, list[str], int]] = []
        
        msg = ''
        if debug_print:
            msg = 'Rows processed: '
        for row in rows:
            try:
                file, town_list = self.get_file_towns_from_cumulative_report(row)
                if file and town_list and (doc_id := self.get_doc_id(file)):
                    pending.append((row, file, town_list, doc_id))
                    if debug_print:
                        msg = f"{msg}{row}, "
            except ValueError as err:
                print(err)
                continue

        if debug_print:
            print(f"{msg}\nA total of {len(pending)}")

        # Flush pending docs in batches
        for flush_start in range(0, len(pending), batch_size):
            chunk = pending[flush_start:flush_start + batch_size]
            buffered_doc_ids = [item[3] for item in chunk]

            # Batch-extract locations for all buffered docs
            if not skip_extraction:
                batch_results = self.file_location_finder.get_doc_locations_batched(
                    buffered_doc_ids,
                    batch_size=batch_size,
                    only_smallest_locations=True,
                    debug_print=debug_print,
                )
            else:
                batch_results = {doc_id: [] for doc_id in buffered_doc_ids}

            # Now evaluate each doc in this chunk using the batched result
            for row, file, town_list, doc_id in chunk:
                print(f"Processing document number {self.num_evaluated_docs} with ID {doc_id}")
                # Swap in the batched result for the extraction step
                self.evaluate_on_one_doc_from_batch(
                    doc_id, town_list, file, batch_results.get(doc_id, []), debug_print
                )

        self.print_statistics(max_num_docs)

    def evaluate_on_one_doc_from_batch(
        self,
        doc_id: int,
        cumulative_towns: list[str],
        file: str,
        doc_location_ids: list[str],
        debug_print: bool,
    ):
        """Like evaluate_on_one_doc but takes doc_location_ids directly (already extracted)."""
        cumulative_locations = self.file_location_finder.match_places_to_location_ids(
            doc_id, cumulative_towns, debug_print)
        cumulative_locations = [dict(f_set) for f_set in cumulative_locations]
        cumulative_locations_ids = [loc["loc_id"] for loc in cumulative_locations]
        cumulative_locations_ids_set = set(cumulative_locations_ids)
        if len(cumulative_locations_ids_set) == 0:
            print(f"None of the towns {cumulative_towns} for document {file} is contained in the JGDB")
            return

        doc_location_ids_set = set(doc_location_ids)
        if doc_location_ids_set == cumulative_locations_ids_set:
            self.num_docs_evaluated_correctly += 1
        if cumulative_locations_ids_set.issubset(doc_location_ids_set):
            self.num_docs_with_all_locations_found += 1

        if debug_print:
            if doc_location_ids_set == cumulative_locations_ids_set:
                msg = f"Record {file}, document ID {doc_id}, all locations coincide: "
            else:
                msg = f"Record {file}, document ID {doc_id}, {len(cumulative_locations_ids_set)} assumed locations: "
                for loc_id in cumulative_locations_ids_set:
                    location = self.file_location_finder.get_location_from_id(str(loc_id))
                    if location:
                        msg = f"{msg} {location['main_name']}"
                msg = f"{msg}; identified {len(doc_location_ids_set)}: "
            for loc_id in doc_location_ids_set:
                location = self.file_location_finder.get_location_from_id(loc_id)
                if location:
                    msg = f"{msg} {location['main_name']}"
            print(msg)

        intersection_set = doc_location_ids_set & cumulative_locations_ids_set

        self.num_evaluated_docs += 1
        self.total_num_cumulative_locs += len(set(cumulative_towns))
        self.total_num_cumulative_locs_in_jgdb += len(cumulative_locations_ids_set)
        self.total_num_extracted_locs += len(doc_location_ids_set)
        self.total_num_coinciding_locs += len(intersection_set)


    def print_statistics(self, max_num_docs:int):
        print(f"Evaluated {self.num_evaluated_docs} documents out of {max_num_docs} - "
              f"{int(0.5 + 100 * self.num_evaluated_docs / max(1, max_num_docs))}%;")
        print(f"{self.num_docs_evaluated_correctly} were identified correctly - "
              f"{int(0.5 + 100 * self.num_docs_evaluated_correctly / max(1, self.num_evaluated_docs))}%")
        print(f"{self.num_docs_with_all_locations_found} had all locations identified - "
            f"{int(0.5 + 100 * self.num_docs_with_all_locations_found / max(1, self.num_evaluated_docs))}%")
        print(f"Total number of locations in the cumulative report for these documents is "
              f"{self.total_num_cumulative_locs}, of these {self.total_num_cumulative_locs_in_jgdb} "
              f"found in the JewishGen DB - "
              f"{int(0.5 + 100 * self.total_num_cumulative_locs_in_jgdb / max(1, self.total_num_cumulative_locs))}%")
        print(f"Number of locations extracted for these documents is {self.total_num_extracted_locs}, of these "
              f"{self.total_num_coinciding_locs} coincide with the cumulative report - locations - "
              f"{int(0.5 + 100 * self.total_num_coinciding_locs / max(1, self.total_num_extracted_locs))}%")


    def get_doc_id(self, file: str):
        possible_labels = self.get_doc_labels(file)

        record_id = None
        # get the first record with one of these labels
        for label in possible_labels:
            cursor = None
            while True:
                records, cursor = self.file_location_finder.scan_database(
                    table_name="Documents",
                    where=("label", "eq", label),
                    limit=100,
                    cursor=cursor,
                )
                if records or cursor is None:
                    break
            if records:
                record_id = records[0]['Id']
                break
        if not record_id:
            print(f"Documents table does not contain a record with the label similar to {file}")
        return record_id

# testing
if __name__ == "__main__":
    report_path = r"C:\Users\user\PycharmProjects\birddog\research\triage\Cumulative Ukraine Research Report.xlsx"
    evaluator = LocationPerformanceEvaluator(report_path)

    debug_print_ = True
    batch_size_ = 5
#    rows_ = [64, 95, 155]
    rows_ = evaluator.unique_random_integers(200)
    evaluator.evaluate_location_extraction(rows_, False, debug_print_, batch_size_)
#    cumulative_file = "ЦДІАК 1167-1-132"
#    print(evaluator.get_doc_id(cumulative_file))
#    file_, town_list_ = evaluator.get_file_towns_from_cumulative_report(5289)
#    if file_ and town_list_ and (doc_id := evaluator.get_doc_id(file_)):
#        evaluator.evaluate_on_one_doc(doc_id, town_list_, file_, False, True)
