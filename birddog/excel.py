# (c) 2025 Jonathan Brandt
# Licensed under the MIT License. See LICENSE file in the project root.


import re
import string
from pathlib import Path
from copy import copy
from openpyxl import load_workbook
from birddog.utility import get_text, ARCHIVE_BASE

import logging
logger = logging.getLogger(__name__)

def is_linked(url):
    return url and not "redlink" in url

def link_status(url):
    return "linked" if is_linked(url) else "unlinked"

def child_url(child):
    link = child[0]["link"]
    return ARCHIVE_BASE + link if link is not None else None

def child_doc_url(parent, child, lru=None):
    #logger.info(f'child_doc_url: child_url {child_url(child)}')
    if not child or not is_linked(child_url(child)):
        #logger.info(f'child_doc_url: {parent.name}: unlinked {get_text(child[0]["text"])}')
        return None
    child_id = get_text(child[0]['text'])
    #logger.info(f'child_doc_url: {parent.name}: looking up {child_id}')
    if lru:
        child = lru.lookup_child(parent, child_id)
    else:
        child = parent[child_id]
    return child.doc_url if child else None

EXPR_PATTERN = re.compile(r'{[^}]+}')

def check_string(text):
    if not text:
        return None
    match = re.findall(EXPR_PATTERN, text)
    if not match:
        return None
    return list(match)

def check_cell(cell):
    return check_string(cell.value)

PARSE_PATTERN = re.compile(
    r'{(?P<expr>[a-zA-Z0-9_.]+)(\[(?P<index>[0-9]+)\])?(\:(?P<modifier>[a-zA-Z_]+))?}')

def parse_template_expr(expr):
    match = re.match(PARSE_PATTERN, expr)
    return match.groupdict() if match else None

def substitute(page, expr):
    try:
        fstring = '{__page__.' + expr['expr'] + '}'
        return string.Formatter().format(fstring, __page__=page)
    except:
        return None

TEMPLATE_DIR = Path(__file__).resolve().parent.parent / 'resources/xlsx_templates'

def export_page(page, dest_file=None, lru=None):
    template_file = f'{TEMPLATE_DIR}/{page.kind}.xlsx'
    logger.info(f"{f'opening template file {template_file}...'}")
    workbook = load_workbook(filename = template_file)
    sheet = workbook.active

    # process title
    title = sheet.title
    check = check_string(title)
    if check:
        for expr in check:
            parsed = parse_template_expr(expr)
            sub = substitute(page, parsed)
            if sub:
                title = title.replace(expr, sub)
    sheet.title = title

    # make a list of all cells with template expressions
    edit_cell = {}
    edits = []
    for row in sheet.iter_rows():
        for cell in row:
            check = check_cell(cell)
            if check:
                parsed = [parse_template_expr(e) for e in check]
                for parse in parsed:
                    # check for formatting directives (before actually editing)
                    if parse['expr'] == 'edit':
                        # {edit} cells contain formatting for editing highlights
                        edit_cell[parse['modifier']] = cell
                edits.append((cell, check, parsed))

    first_child_row = None
    last_child_row = None
    columns = []

    for edit in edits:
        cell, matches, parses = edit
        for match, parse in zip(matches, parses):
            if parse['expr'] == 'col':
                # defer {col} expressions until after other substitutions are done
                # since this fills in the table and will overwrite the {rollup} expressions
                first_child_row = row = cell.row
                last_child_row = first_child_row + len(page.children) - 1
                columns.append((cell, parse))
            elif parse['expr'] == 'rollup':
                # generate a formula based on rollup expression
                rollup_cell = sheet.cell(row=cell.row - 1 + len(page.children), column=cell.column)
                col = cell.column_letter
                rollup_cell.value = f'={parse["modifier"]}({col}{first_child_row}:{col}{last_child_row})'
                rollup_cell.style = cell.style
                rollup_cell.border = copy(cell.border)
                rollup_cell.alignment = copy(cell.alignment)
                rollup_cell.font = copy(cell.font)
                cell.value = ""
            elif parse['expr'] == 'edit':
                # {edit} cells contain formatting for editing highlights (caught above)
                pass
            else:
                # general case: replace template expression with substitution value
                sub = substitute(page, parse)
                if sub is not None:
                    cell.value = cell.value.replace(match, sub)
            if parse['modifier'] == 'linked':
                cell.hyperlink = page.url

    # now expand each of the {col...} expressions
    for cell, parse in columns:
        row = cell.row
        col = cell.column
        index = parse['index']
        for child in page.children:
            child_cell = sheet.cell(row=row, column=col)
            if child_cell.row != cell.row:
                # propagate border, style, font, and alignment to all rows
                child_cell.style = cell.style
                child_cell.border = copy(cell.border)
                child_cell.alignment = copy(cell.alignment)
                child_cell.font = copy(cell.font)
            if index is not None:
                item = child[int(index)]
                sub = get_text(item['text'])
                if 'edit' in item:
                    edit = item['edit']
                    if edit in edit_cell:
                        child_cell.fill = copy(edit_cell[edit].fill)
                child_cell.value = sub
            else:
                child_cell.value = ''
            if parse['modifier'] == 'linked':
                child_cell.hyperlink = child_url(child)
            elif parse['modifier'] == 'doc_link':
                child_cell.hyperlink = child_doc_url(page, child, lru)
            elif parse['modifier'] == 'link_status':
                child_cell.value = link_status(child_url(child))
            row += 1

    if dest_file:
        workbook.save(dest_file)
    return workbook