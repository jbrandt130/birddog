from birddog.abstract_database import InvalidFieldValue
from birddog.database import Database, timer
from birddog.database_updater import form_document_record, normalize_url
from birddog.log import get_logger
from birddog.utility import utc_now_dt
from birddog.wiki import (
    WIKI_NAMESPACE,
    get_title,
    parent_title,
    page_label,
    sequential_page_label,
)

from datetime import datetime
from openpyxl import load_workbook
from pathlib import Path
from typing import List, Any, Literal
from urllib.parse import urlparse, parse_qs, urlunparse, unquote, quote
import os
import re
import time

_ENABLE_CONSOLE_HIGHLIGHTING = os.name == "nt"  # console output highlighting
START_HDR_ROW = 5  #the headers are sometimes in row 5 and sometimes in row 6 or 7
END_HDR_ROW = 7
FIRST_OPUS_HDR1 = "Case #, link".upper()
FIRST_OPUS_HDR2 = "P#, link".upper()
FIRST_OPUS_HDR3 = "Index #, link".upper()
FIRST_OPUS_HDR4 = "NEW NUMBERING".upper()
FIRST_OPUS_HDR5 = "#".upper()
SECOND_OPUS_HDR = "Case description, file link".upper()
FIRST_FUND_HDR = "Opus #, link".upper()
SECOND_VOLUME_HDR = "Other case #".upper()
AVAILABILITY_HDR = "Availability".upper()
SRC_HDR = "source:".upper()
SRC_NO_COLON_HDR = "source".upper()
COMMENTS_HDR = "Comments".upper()
CHANGE_DATE_HDR = "change date:".upper()
REF_DATE_HDR = "reference date:".upper()
RED_SUBSTRING = "action=edit&redlink=1"  # it can be either &action=edit&redlink=1 or ?action=edit&redlink=1 in the end
_NS_PREFIX = f"{WIKI_NAMESPACE}:"

DESCRIPTION_COL_OFFS = 1
YEARS_COL_OFFS = 2
DOC_TYPE_COL_OFFS = 3
CONTENT_CODE_COL_OFFS = 4
PROCESS_CODE_COL_OFFS = 5

if _ENABLE_CONSOLE_HIGHLIGHTING:
    from colorama import Fore, init

    init()  # Windows fix
else:
    # stub out highlight directives
    class Fore:
        RESET = ""
        RED = ""
        GREEN = ""

_logger = get_logger()


def parent_title_no_ns(title, wiki_spreadsheet, archive_unit_name):
    if wiki_spreadsheet:
        result = parent_title(title)
        if not result:
            return ""
        if result.startswith(_NS_PREFIX):
            result = result[len(_NS_PREFIX):]
        return result
    else:
        last_slash_pos = archive_unit_name.rfind('/')
        if last_slash_pos != -1:
            return archive_unit_name[0: last_slash_pos]
        else:
            raise ValueError(f"Archive unit name: {archive_unit_name} contains no slashes")


def get_case_id(cell_contents) -> str:
    """If the cell contents is a string representing a number,
    we return its integer part. Otherwise, we return the original string."""
    try:
        case_id_float = float(cell_contents)
        case_id = str(int(case_id_float))
        return case_id
    except ValueError:
        return cell_contents


def canonical_wiki_url(url: str) -> str:
    """Convert a MediaWiki index.php?title= URL to /wiki/Title style."""
    parsed = urlparse(url)
    if parsed.path != "/w/index.php" and not parsed.path.endswith("/w/index.php"):
        return url  # leave non‑index.php URLs unchanged

    qs = parse_qs(parsed.query)
    if not qs.get("title"):
        return url

    title = qs["title"][0]
    # For wikisource‑style /wiki/… paths (not /w/…)
    nice_path = f"/wiki/{title}"
    # quote path component if needed (e.g. non‑ASCII)
    nice_path = "/".join(quote(p, safe="") for p in nice_path.split("/"))

    url = urlunparse((
        parsed.scheme,
        parsed.netloc,
        nice_path,
        parsed.params,
        "",  # no query
        parsed.fragment,
    ))
    url = normalize_url(url)
    return url


def url_to_title(url: str) -> str:
    if "index.php" in url:
        query = urlparse(url).query
        params = parse_qs(query)
        if "title" in params:
            result = params["title"][0]
            ns_prefix = f"{WIKI_NAMESPACE}:"
            if result.startswith(ns_prefix):
                result = result[len(ns_prefix):]
            return result
    if "redlink" in url:
        url = urlparse(url).path
    return get_title(unquote(url), include_namespace=False)


def cell_to_title(cell, archive_unit_name) -> str:
    val = cell.value
    if isinstance(val, float):
        subunit_name = str(int(val))
    else:
        subunit_name = str(val).strip()
    return f"{archive_unit_name}/{subunit_name}"


def get_page_title_from_link(cell, wiki_spreadsheet, archive_unit_name, latin_archive_unit_name):
    if wiki_spreadsheet:
        if cell.hyperlink:
            url = cell.hyperlink.target.strip()
        else:
            if isinstance(cell.value, str):
                url = cell.value
            else:
                try:
                    url = str(int(float(cell.value)))
                    return url, url
                except ValueError:
                    return None, None
        latin_title = latin_archive_unit_name
        if 'wiki' in url:
            title = url_to_title(url)
        else:
            #for URLs like https://forum.j-roots.info/viewtopic.php?p=106843/5/1
            title = archive_unit_name
            try:
                int_contents = str(int(float(cell.value)))
                title = f"{title}/{int_contents}"
                latin_title = f"{latin_archive_unit_name}/{int_contents}"
            except ValueError:
                pass

        return title, latin_title
    else:
        title = cell_to_title(cell, archive_unit_name)
        latin_title = cell_to_title(cell, latin_archive_unit_name)
        return title, latin_title


def parse_hyphen_triplet(s: str):
    parts = s.split('-')
    if len(parts) != 3:
        return [], False
    try:
        nums = [int(p) for p in parts]
    except ValueError:
        return [], False
    return nums, True


def get_case_title(url, curr_parent_title, opus_name, cell, wiki_spreadsheet,
                   possible_hyphen_in_case_num, archive_unit_name, archive_unit_cyrillic_name):
    if wiki_spreadsheet:
        if possible_hyphen_in_case_num:
            # sometimes instead of a case number we have (fund)-(opus)-(case)
            three_numbers, is_triplet = parse_hyphen_triplet(cell.value)
            if is_triplet:
                title = f"{curr_parent_title}/{three_numbers[1]}/{three_numbers[2]}"
                return title, title

        title = url_to_title(url)
        if curr_parent_title not in title:
            value = cell.value
            if isinstance(value, float) and value.is_integer() or isinstance(value, int):
                case_id = get_case_id(value)
            else:
                case_id = value
            title = f"{curr_parent_title}/{opus_name}/{case_id}"

        return title, title

    else:
        title = cell_to_title(cell, archive_unit_cyrillic_name)
        latin_title = cell_to_title(cell, archive_unit_name)
        return latin_title, title


def get_cell_link(cell):
    return unquote(cell.hyperlink.target.strip()) if cell.hyperlink else ""


def add_to_message(message, addition):
    if message is None or message.strip() == "":
        message = addition
    else:
        message = message + "; " + addition
    return message


def string_contains_part(link, necessary_part, strict_equality):
    # does the link contain the necessary part?
    contains = necessary_part in link
    if not contains:
        #replaces all the hyphens and all the underscores in the link with a slash
        modified_link = link.replace('-', '/').replace('_', '/')
        contains = necessary_part in modified_link

    necessary_part1 = re.sub(r"[^0-9/]+", "", necessary_part)
    if not contains:
        # try without letters
        contains = necessary_part1 in link
        
    if not contains:
        # try space and slash separated substrings
        subparts = re.split(r"[ /-]+", necessary_part)
        contains = True
        for subpart in subparts:
            contains = contains and re.sub(r'[^0-9/]+', '', subpart) in link

    if not contains and not strict_equality:
        #try without zeros
        link_no_zeros = link.replace('0', '')  # Removes all '0's
        contains = necessary_part1 in link_no_zeros
        if not contains:
            necessary_part1 = necessary_part1.replace('0', '')  # Removes all '0's
            contains = necessary_part1 in link_no_zeros
            if not contains:
                #try replacing '-' with slash
                link_no_zeros = link_no_zeros.replace('-', '/')
                contains = necessary_part1 in link_no_zeros

    return contains


def consistent_link(cell, cell_address, archive_unit_name, sheet_title, wiki_spreadsheet, import_message):
    """Checks whether the cell contents and the link coincide,
    and that the address contains the fund and opus name when relevant"""
    slash_post = archive_unit_name.find('/')
    necessary_part = archive_unit_name[slash_post + 1:] if slash_post != -1 else ''

    contents = unquote(cell.value.strip())
    mess = ""  # default
    link = ""  # default
    if cell.hyperlink:
        link = unquote(cell.hyperlink.target.strip())
        coincides = link == contents
        link_contains_necessary_part = string_contains_part(link, necessary_part, wiki_spreadsheet)
        if coincides:
            if not link_contains_necessary_part and wiki_spreadsheet:
                mess = (f"cell {cell_address} in worksheet {sheet_title} contents and link '{contents}'"
                        f" were supposed to include '{necessary_part}'")
        else:
            #contents vs link
            if '/forum.j-roots.info' in link:
                mess = (f"the link '{link}' in cell {cell_address}, sheet {sheet_title},"
                        f" differs from its contents '{contents}'")
            elif link_contains_necessary_part:
                mess = (f"using the link '{link}' in cell {cell_address}, sheet {sheet_title},"
                        f" that differs from its contents '{contents}'")
            elif string_contains_part(contents, necessary_part, wiki_spreadsheet):
                mess = (f"using the contents '{contents}' in cell {cell_address}, sheet {sheet_title},"
                        f" instead of the apparently wrong link '{link}'")
                link = contents
            else:
                mess = f"the link '{link}' in cell {cell_address} in sheet {sheet_title} is apparently wrong"
                if wiki_spreadsheet:
                    raise TypeError(mess)
    elif contents != "":
        if contents.upper().startswith('HTTP'):
            link = contents
            mess = f"cell {cell_address} in sheet {sheet_title} does not have a link, using instead its contents '{contents}'"
        else:
            mess = f"cell {cell_address} in sheet {sheet_title} does not have a link and contains text '{contents}'"
    else:
        mess = f"cell {cell_address} in sheet {sheet_title} does not have a link"
        raise TypeError(mess)

    import_message = add_to_message(import_message, mess)
    return link, import_message


def get_cell_value(cell, capitalized=False):
    value = cell.value
    if value is None:
        return value
    if isinstance(value, float):
        value = int(value)
    if not isinstance(value, str):
        value = str(value)
    value = value.strip()
    if capitalized:
        value = value.upper()

    return value


def parse_http_list(s: str):
    # separators - comma, semicolon, and "and" with optional spaces before/after
    # we exclude commas that are followed by an underscore, because that is what we have in cell 'B4', sheet 'fund 315',
    # spreadsheet DAHMO-D-wiki-20260415.xlsx
    # we also exclude commas followed by optional spaces and a Cyrillic character,
    # because that is what we have in cell 'B4', sheet 'fund 1517', spreadsheet DAKO-D-general-wiki-20260417.xlsx
    parts = re.split(r"\s*and\s*|,(?!_|[\t ]*[\u0400-\u04FF])|;", s)
    #parts = re.split(r"\s*and\s*|,\s+(?=http)|;", s, flags=re.IGNORECASE)
    #parts = re.split(r"\s*and\s*|,(?![\t ]*_)|;", s)
    #parts = re.split(r"\s*and\s*|,(?![\t ]*[\u0400-\u04FF])|;", s)
    #parts = re.split(r"\s*and\s*|,(?!_)|;", s)
    # Filter out empty strings (from consecutive separators)
    parts = [p.strip() for p in parts if p.strip()]
    parts = [canonical_wiki_url(p.strip()) for p in parts]
    if len(parts) > 1 and all(part.upper().startswith('HTTP') for part in parts):
        return parts
    return [canonical_wiki_url(s)]


def get_cell_link_or_str(ws, cell_addr: str, import_message: str, check_contents: bool = True):
    cell = ws[cell_addr]
    result = get_cell_link(cell)
    mess = ""
    used_contents = False
    if result == "" and check_contents:
        result = get_cell_value(cell)
        used_contents = True
    if result:
        result = parse_http_list(result)
        if len(result) > 1:
            _logger.warning(f"{len(result)} links in cell {cell_addr}")
        elif result[0] and used_contents and result[0].upper().startswith('HTTP') and '4' not in cell_addr:
            #Oxana asks not to issue this message for links in cell B4
            mess = f"no document link in cell {cell_addr}, using the cell contents '{result[0]}'"

        for res in result:
            looks_like_a_link = res.upper().startswith('HTTP') and string_ends_with_file_ext(res)
            if not looks_like_a_link and ' ' in res:
                mess = f"Document link '{res}' in cell '{cell_addr}', sheet '{ws.title}' may contain comments"
    if mess:
        _logger.warning(mess)
        import_message = add_to_message(import_message, mess)
    return import_message, result

def string_ends_with_file_ext(s: str):
    s = s.upper()
    return s.endswith(".PDF") or s.endswith(".DJVU") or s.endswith(".ZIP")


def get_doc_links(ws, cell_addr: str, import_message: str, archive_unit_link: str, only_check_link: bool):
    modified_import_message, cell_links = get_cell_link_or_str(ws, cell_addr, import_message, False)
    if not cell_links:
        return import_message, ""
    ar1 = 'Архіви/'
    ar2 = 'Архів:'
    # we compare the strings up to replacement of ar1 with ar2
    archive_unit_link_replaced = canonical_wiki_url(archive_unit_link.rstrip(" /").replace(ar1, ar2))
    if isinstance(cell_links, str) and cell_links.replace(ar1, ar2) == archive_unit_link_replaced:
        if only_check_link:
            cell_links = ""
        return import_message, cell_links

    real_doc = True
    if isinstance(cell_links, list) and len(cell_links) == 1:
        cell_link_replaced = cell_links[0].rstrip(" /").replace(ar1, ar2)
        if cell_link_replaced in archive_unit_link_replaced or archive_unit_link_replaced in cell_link_replaced:
            if only_check_link:
                cell_links = ""
            return import_message, cell_links
        real_doc = string_ends_with_file_ext(cell_link_replaced)

    mess = ''
    if not real_doc and only_check_link:
        mess = (f" Link {cell_links} in cell '{cell_addr}', sheet '{ws.title}' "
                f"differs from the page url {archive_unit_link}")
        cell_links = ""

    if len(mess) > 0:
        modified_import_message = add_to_message(modified_import_message, mess)
        _logger.warning(mess)
    if real_doc and only_check_link:
        _logger.warning(f"Adding document links {cell_links} found in cell '{cell_addr}', sheet '{ws.title}'")

    return modified_import_message, cell_links


def get_cell_int_value(cell):
    value = get_cell_value(cell)
    if value is None:
        return 0
    return int(value)


def combine_cell_value(cell1, cell2):
    result = ",".join(filter(
        lambda x: x is not None,
        (get_cell_value(cell1), get_cell_value(cell2))))
    return result if result else None


def is_url_redlinked(url):
    is_redlinked = False  #default
    naked_url = url  #default
    index = url.find(RED_SUBSTRING)
    if index != -1:
        is_redlinked = True
        # it can be either &action=edit&redlink=1 or ?action=edit&redlink=1 in the end
        naked_url = url[:index - 1]

    return is_redlinked, naked_url


def check_redlink(page: dict) -> dict:
    availability = page.get("availability", "linked")
    url = page["url"]
    is_redlinked, naked_url = is_url_redlinked(url)
    if is_redlinked:
        # if the URL contains a red_substring, availability status is "redlinked"
        url = naked_url
        availability = "redlinked"
        page["url"] = url
    page["availability"] = availability

    return page


def delete_part_between_hyphen_and_slash(s: str) -> str:
    # Get part before first slash
    if "/" not in s:
        part_before_slash = s
        part_after_slash = ''
    else:
        part_before_slash = s.split("/", 1)[0]
        part_after_slash = '/' + s.split("/", 1)[1]

    # Get part before first hyphen from the part before the slash
    if "-" not in part_before_slash:
        result = s
    else:
        part_before_hyphen = part_before_slash.split("-", 1)[0]
        result = part_before_hyphen + part_after_slash

    return result


def add_page(page: dict, page_table: dict) -> None:
    """Merge a page entry into the page_table by title."""
    #(re)set some fields
    page = check_redlink(page)
    url = page["url"]
    url = canonical_wiki_url(url)
    page["url"] = url
    page["last_imported"] = str(utc_now_dt().replace(microsecond=0))
    page["label"] = delete_part_between_hyphen_and_slash(page["label"])

    # ensure an entry exists for this URL
    entry = page_table.setdefault(url, dict())
    # update/merge keys
    entry.update(page)


def normalize_to_int_str(s: str) -> str:
    if not isinstance(s, str):
        return s

    # Check if integer
    if re.fullmatch(r'-?\d+', s):
        return str(int(s))

    # Check if float that converts to integer
    if re.fullmatch(r'-?\d+(\.\d+)?', s):
        try:
            f = float(s)
            if f.is_integer():
                return str(int(f))
        except ValueError:
            pass

    return s

def irregular_url(url: str) -> bool:
    # URLs in these domains do not contain fond number etc
    return ('drive.google.com' in url or 'alexkrakovsky.direct.quickconnect.to' in url or
            'gov.ua/files/' in url or 'forum.j-roots.info' in url or '.gov.ua/inventories/' in url or
            'gov.ua/?page_id=' in url)


def log_strange_parsing_result(ws, r, cell, subunits: List[str], url: str, wiki_spreadsheet: bool, level: str):
    match level:
        case "opus":
            lower_level = "case"
        case "archive":
            lower_level = "fund"
        case _:
            message = f"Illegal level {level}"
            _logger.error(message)
            raise ValueError(message)

    if irregular_url(url):
        # URLs in these domains do not contain fond number etc
        return lower_level

    num_results = len(subunits)
    normalized_val = normalize_to_int_str(str(cell.value))

    match num_results:
        case 1:
            #default result - check url sanity
            if wiki_spreadsheet and normalized_val not in url:
                return lower_level
                no_letters = re.sub(r"[A-Z-]", "", normalized_val)
                if no_letters not in url:
                    _logger.warning(f"In sheet='{ws.title}', row={r}, the URL {url} was supposed to include"
                                    f" the {lower_level} number '{normalized_val}'")

        case 0:
            if cell.value and len(str(cell.value)) > 0:
                #otherwise, it is an empty cell - do nothing
                _logger.warning(
                    f"No URL found for the {lower_level} '{normalized_val}', sheet='{ws.title}', row={r} - skipping")

        case _:
            _logger.warning(f"Irregular {lower_level} number '{normalized_val}', sheet='{ws.title}', "
                     f"row={r} - splitting to {subunits}")

    return lower_level


def get_url_from_2_cells(unit_num_cell, unit_descr_cell) -> tuple[str, bool]:
    cell_to_use = unit_num_cell
    in_first_cell = True
    if cell_to_use.hyperlink is None:
        cell_to_use = unit_descr_cell
        in_first_cell = False
    if cell_to_use.hyperlink is None:
        return "", in_first_cell

    hyperlink = cell_to_use.hyperlink
    url = unquote(hyperlink.target.strip())
    url = canonical_wiki_url(url)

#    if hyperlink.location is not None:
#        url = f"{url}#{hyperlink.location}"
    return url, in_first_cell


def parse_cell_integers(fund_num_cell, fund_descr_cell, archive_latin_name) -> tuple[list[Any], str, bool]:
    """
    Parses an openpyxl cell value into a list of integers based on specific rules.
    If no link found in both fund_num_cell, fund_descr_cell, the resulting list is empty


    Rules:
    1. Float that is integer (e.g., 41.0) -> [41]; integer -> integer
    2. String with comma-separated integers (e.g., "41, 42") -> [41, 42]
    3. datetime object -> [day, month]
    4. String range "10-13" -> [10, 11, 12, 13]
    5. Otherwise -> []
    All the integers are converted to strings.
    For IMFE archive, the fund ID includes hyphens, like 14-8.

    Args:
        fund_num_cell: openpyxl.cell.cell.Cell object
        fund_descr_cell: openpyxl.cell.cell.Cell object
        archive_latin_name: Archive name in Latin letters, used for archive-specific parsing rules.

    Returns:
        1) List[str]: List of parsed integers converted to strings
        2) str: the url from one of the two input cells.
    """
    value = fund_num_cell.value
    if value is None:
        return [], "", False

    url, in_first_cell = get_url_from_2_cells(fund_num_cell, fund_descr_cell)
    if url == "":
        return [], "", in_first_cell

    # Case 1: Float that is integer
    if isinstance(value, float) and value.is_integer():
        return [str(int(value))], url, in_first_cell

    # Case 1: integer
    if isinstance(value, int):
        return [str(value)], url, in_first_cell

    # Case 2 & 4: Convert to string and parse
    s = str(value).strip()

    if '-' in s:
        if archive_latin_name == 'IMFE':
            return [s], url, in_first_cell

        if re.match(r'^\d+-\d+$', s):
            # Case 4: Range like "10-13"
            try:
                start, end = map(int, s.split('-'))
                return [str(x) for x in list(range(start, end + 1))], url, in_first_cell
            except ValueError:
                pass
        pattern = r'^[A-Z]-[1-9]\d*$'
        is_prefixed_number = bool(re.match(pattern, s))
        if is_prefixed_number:
            return [s], url, in_first_cell

    # Case 2: Comma-separated "41, 42"
    if ',' in s:
        parts = [p.strip() for p in s.split(',')]
        ints = []
        for p in parts:
            try:
                num = float(p)
                if num.is_integer():
                    ints.append(str(int(num)))
            except ValueError:
                pass
        if ints:
            return ints, url, in_first_cell

    # Case 4: datetime -> [day, month]
    if isinstance(value, datetime):
        return [str(value.day), str(value.month)], url, in_first_cell

    if not in_first_cell:
        return [s], url, in_first_cell

    return [], "", in_first_cell


def is_series_of_case_numbers(s: str, ignore_parts: bool) -> tuple[bool, int, bool]:
    """Case number should include digits but may also include letters.
    Case numbers are separated by hyphens."""
    s = s.strip()
    inconsistent_parts = False
    if not isinstance(s, str) or not s:
        return False, 0, inconsistent_parts

    parts = s.split('-')
    num_parts = len(parts)
    if any(part == '' for part in parts):
        return False, num_parts, inconsistent_parts

    only_integers_found = False
    integers_and_letters_found = False
    for part in parts:
        if re.fullmatch(r'\d+', part):
            only_integers_found = True
            continue
        if re.fullmatch(r'\d+\.\d+', part) and float(part).is_integer():
            only_integers_found = True
            continue
        if any(ch.isdigit() for ch in part):
            integers_and_letters_found = True
            continue
        return False, num_parts, inconsistent_parts

    if ignore_parts:
        return True, 1, False

    inconsistent_parts = only_integers_found and integers_and_letters_found
    if inconsistent_parts:
        #DAOO-D-wiki-20260413 sheet DAOO 1-191 contains case number 1-T2
        return True, 1, inconsistent_parts

    return True, num_parts, inconsistent_parts


def find_header_in_row(ws, row, start_col, end_col, header):
    """Find a cell with the given header in the given row of the given worksheet.
    Returns the column if found, None otherwise."""
    start_ord = ord(start_col)
    end_ord = ord(end_col)
    for i in range(start_ord, end_ord + 1):
        cell_addr = f"{chr(i)}{row}"
        contents = get_cell_value(ws[cell_addr], True)
        if contents is None:
            continue
        if header in contents:
            return i
    return None


def find_header_in_col(ws, start_row, end_row, col, headers):
    """Find a cell with one of the given headers in the given column of the given worksheet.
    Returns 1) the row if found, None otherwise.
            2) The header found """
    for header in headers:
        for i in range(start_row, end_row + 1):
            cell_addr = f"{col}{i}"
            contents = get_cell_value(ws[cell_addr], True)
            if contents is None:
                continue
            if contents == header:
                return i, header
            if header == FIRST_OPUS_HDR4 and header in contents:
                # it can be something like "NEW NUMBERING (listed here 06/2025)"
                return i, header
    return None, None


def get_source_type(ws):
    #it is usually in C1, but it may occur also in E1
    start_ord = ord("C")
    end_ord = ord("E")
    source_type = None  #default
    for i in range(start_ord, end_ord + 1):
        cell_addr = f"{chr(i)}1"
        source_type = get_cell_value(ws[cell_addr])
        if source_type is None:
            continue

        #it is sometimes "archive" instead of "archives"
        if source_type.startswith("archive"):
            source_type = "archives"
        elif source_type.startswith("wiki"):
            source_type = "wiki"
        else:
            source_type = "other"
        break
    return source_type


def get_dates(ws, archive_name, change_date_col, ref_date_col):
    change_date = get_cell_value(ws[f"{chr(change_date_col + 1)}1"])  # default L1
    timestamp = get_cell_value(ws[f"{chr(ref_date_col + 1)}1"])  # default 01
    if str(timestamp) == '?' and "DAOO" in archive_name:
        # Juliana's recommendation for these specific spreadsheets
        timestamp = "09 Nov 2024"
        _logger.warning(f"Using change date of {timestamp}")

    return change_date, timestamp


def get_url_and_parent_title(ws, wiki_spreadsheet, archive_unit_name, archive_cyrillic_unit_name):
    sheet_title = ws.title
    import_message = ""
    row = 3
    source_col = find_header_in_row(ws, row, "B", "Z", SRC_HDR)
    daho_fund_list = False
    if source_col is None:  # and 'DAHO' in archive_unit_name and not wiki_spreadsheet:
        # try SRC_NO_COLON_HDR
        source_col = find_header_in_row(ws, row, "B", "Z", SRC_NO_COLON_HDR)
        if source_col is None:
            # non-standard spreadsheet 'DAHO'
            daho_fund_list = True
            row = 4
            source_col = ord('A')
        else:
            cell_address = f"{chr(source_col)}{row}"  # A4
            mess = ws[cell_address].value.strip()
            import_message = add_to_message(import_message,
                                            mess[:len(mess) - 1])  # deleting the confusing colon in the end

    cell_address = f"{chr(source_col + 1)}{row}"  # default D3
    parent_title_cell = ws[cell_address]
    url, import_message = consistent_link(parent_title_cell, cell_address, archive_unit_name,
                                          sheet_title, wiki_spreadsheet, import_message)
    if daho_fund_list:
        # additional links
        cell_address = f"{chr(source_col)}{row}"  # A4
        descr_cell = ws[cell_address]
        import_message = add_to_message(import_message, f"The main link is marked as {descr_cell.value.strip()}")
        cell_address = f"{chr(source_col)}{row + 1}"  # A5
        descr_cell = ws[cell_address]
        cell_address = f"{chr(source_col + 1)}{row + 1}"  # B5
        link_cell = ws[cell_address]
        import_message = add_to_message(import_message, f"there is also a {descr_cell.value.strip()} link "
                                                        f"{unquote(link_cell.hyperlink.target.strip())}")
        descr_cell = ws["C3"]
        link_cell = ws["D3"]
        import_message = add_to_message(import_message, f", a {descr_cell.value.strip()} link "
                                                        f"{unquote(link_cell.hyperlink.target.strip())}")
        descr_cell = ws["C4"]
        link_cell = ws["D4"]
        import_message = add_to_message(import_message, f", a {descr_cell.value.strip()} link "
                                                        f"{unquote(link_cell.hyperlink.target.strip())}")
        descr_cell = ws["C3"]
        link_cell = ws["D3"]
        import_message = add_to_message(import_message, f", and a {descr_cell.value.strip()} link "
                                                        f"{link_cell.value}")

    if wiki_spreadsheet:
        title, latin_title = get_page_title_from_link(parent_title_cell, wiki_spreadsheet,
                                                      archive_cyrillic_unit_name, archive_unit_name)
    else:
        title = archive_cyrillic_unit_name
        latin_title = archive_unit_name

    if import_message != "":
        _logger.warning(import_message)

    #comments can be either in cell I2 or I3
    comments = get_cell_value(ws["I2"])
    if not comments:
        comments = get_cell_value(ws["I3"])
#    if comments:
#        _logger.info(f"There is a comment: {comments}")

    return title, latin_title, url, import_message, comments


def to_positive_int_str(s: str) -> str:
    """
    If input converts to positive integer (>0), returns str(int).
    Otherwise, returns original input unchanged.
    Handles "11.0" → "11", rejects "0", negatives, non-numbers.
    """
    stripped = s.strip()
    if not stripped:
        return s
    try:
        num = float(stripped)
        if num.is_integer() and num > 0:
            return str(int(num))
        else:
            return s
    except ValueError:
        return s


def archive_name_to_cyrillic(latin_name: str) -> str:
    match latin_name:
        case "CDIAK":
            return "ЦДІАК"
        case "DADO":
            return "ДАДО"
        case "DAHEO-D":
            return "ДАХЕО-Д"
        case "DAHEO-R":
            return "ДАХЕО-Р"
        case "DAHO":
            return "ДАХО"
        case "DAK":
            return "ДАК"
        case "DAMO-A":
            return "ДАМО-А"
        case "DAMO-D":
            return "ДАМО-Д"
        case "DAMO-Soviet":
            return "ДАМО"
        case "DAOO":
            return "ДАОО"
        case "DAVO":
            return "ДАВО"
        case "TSDAHOU":
            return "ЦДАГОУ"
        case "TSDAVO":
            return "ЦДАВО"
        case _:
            message = f"No Cyrillic equivalent for archive name {latin_name}"
            _logger.error(message)
            raise ValueError(message)


def cyrillic_to_latin(s: str) -> str:
    """
    Converts Ukrainian/Russian Cyrillic letters to Latin equivalents.
    Leaves digits, hyphens, punctuation unchanged.
    """
    # Cyrillic → Latin transliteration map (Russian + Ukrainian)
    trans_map = {
        # Russian lowercase
        'а': 'a', 'б': 'b', 'в': 'v', 'г': 'g', 'д': 'd', 'е': 'e', 'ё': 'yo',
        'ж': 'zh', 'з': 'z', 'и': 'i', 'й': 'y', 'к': 'k', 'л': 'l', 'м': 'm',
        'н': 'n', 'о': 'o', 'п': 'p', 'р': 'r', 'с': 's', 'т': 't', 'у': 'u',
        'ф': 'f', 'х': 'h', 'ц': 'ts', 'ч': 'ch', 'ш': 'sh', 'щ': 'shch',
        'ъ': "'", 'ы': 'y', 'ь': "'", 'э': 'e', 'ю': 'yu', 'я': 'ya',

        # Ukrainian lowercase (+ unique letters)
        'ґ': 'g', 'є': 'ye', 'і': 'i', 'ї': 'yi',

        # Uppercase (same transliteration, capitalized)
        'А': 'A', 'Б': 'B', 'В': 'V', 'Г': 'G', 'Д': 'D', 'Е': 'E', 'Ё': 'Yo',
        'Ж': 'Zh', 'З': 'Z', 'И': 'I', 'Й': 'Y', 'К': 'K', 'Л': 'L', 'М': 'M',
        'Н': 'N', 'О': 'O', 'П': 'P', 'Р': 'R', 'С': 'S', 'Т': 'T', 'У': 'U',
        'Ф': 'F', 'Х': 'H', 'Ц': 'Ts', 'Ч': 'Ch', 'Ш': 'Sh', 'Щ': 'Shch',
        'Ъ': "'", 'Ы': 'Y', 'Ь': "'", 'Э': 'E', 'Ю': 'Yu', 'Я': 'Ya',
        'Ґ': 'G', 'Є': 'Ye', 'І': 'I', 'Ї': 'Yi'
    }

    result = []
    for char in s:
        result.append(trans_map.get(char, char))  # Keep non-Cyrillic chars unchanged

    return ''.join(result)


def title_cell_val_identical(title: str, cell_val: str, num_parts: int, inconsistent_parts: bool) -> tuple[bool, str]:
    """Returns everything after the last '/' or original string if no slash."""
    title_after_last_slash = title.rsplit('/', 1)[-1] if '/' in title else title
    title_after_last_slash = cyrillic_to_latin(title_after_last_slash)
    cell_val_posit_int_cyr = to_positive_int_str(cell_val)
    cell_val_posit_int_lat = cyrillic_to_latin(cell_val_posit_int_cyr)
    identical = title_after_last_slash == cell_val_posit_int_lat
    if not identical:
        # try deleting the trailing non-digits
        identical = re.sub(r'\D*$', '', title_after_last_slash) == re.sub(r'\D*$', '', cell_val_posit_int_lat)
    if not identical and num_parts == 3:
        #maybe it is of format fund-opus-case
        parts = cell_val.split('-')
        all_parts_found = True
        for part in parts:
            if -1 == title.find(part):
                all_parts_found = False
                break
        if all_parts_found:
            identical = True

    modified_title = title
    if not identical and not inconsistent_parts:
        modified_title = title.rsplit('/', 1)[0] + "/" + cell_val_posit_int_cyr if '/' in title \
            else cell_val_posit_int_cyr
    return identical, modified_title


def general_page_label(latin_title, cyrillic_title, wiki_spreadsheet):
    if wiki_spreadsheet:
        return page_label(cyrillic_title)
    else:
        return latin_title


def check_url_sanity(url, import_message, wiki_spreadsheet, necessary_parts, sheet_title, cell_address):
    if irregular_url(url):
        # URLs in some domains do not contain fond number etc
        return import_message

    mess = ""
    if wiki_spreadsheet:
        sequence = ""
        for part in necessary_parts:
            if part:
                sequence = f"{sequence}/{part}"
        sequence = re.sub(r'^\D*', '', sequence)
        if not string_contains_part(url, sequence, wiki_spreadsheet):
            mess = (f"url {url} in cell {cell_address} in worksheet '{sheet_title}'"
                    f" was supposed to include '{sequence}'")
#    else:
#        for part in necessary_parts:
#            if not string_contains_part(url, part, wiki_spreadsheet):
#                mess = (f"url {url} in cell {cell_address} in worksheet '{sheet_title}'"
#                       f" was supposed to include '{part}'")
#                break

    if mess != "":
        _logger.warning(mess)
        import_message = add_to_message(import_message, mess)

    return import_message


def get_language(ws, cell_address):
    language = get_cell_value(ws[cell_address], False)
    if language:
        language = re.split(r"[;, ]+", language)
        _logger.warning(f"Found language {language}")
    return language


def count_fund_opus_attributes(sheet, end_col):
    """Fund name and opus name appear in the first row,
    typically but not always in cells G1 and H1. We count the number of nonempty cells.
    For archive, there would be none, for fund - one, anf for opus - two."""
    start_col = "D"
    row = 1
    start_ord = ord(start_col)
    num_nonempty_cells = 0
    fund_opus_attributes = []
    for i in range(start_ord, end_col):
        cell_addr = f"{chr(i)}{row}"
        contents = get_cell_value(sheet[cell_addr])
        if contents is not None and any(ch.isdigit() for ch in contents):
            #in DAHO-D-wiki-20260324.xlsx sheet "DAHO D wiki fund list" we have text
            # "to be updated when BirdDog not adding incorrect links"
            num_nonempty_cells += 1
            fund_opus_attributes.append(contents)
    return num_nonempty_cells, fund_opus_attributes


def sheet_contains_volumes(ws):
    #the header row can be 5, 6 or 7
    headers = [FIRST_FUND_HDR]
    case_num_col = "A"
    hdr_row, _ = find_header_in_col(ws, START_HDR_ROW, END_HDR_ROW, case_num_col, headers)
    if hdr_row is None:
        case_num_col = chr(ord(case_num_col) + 1)
        hdr_row, _ = find_header_in_col(ws, START_HDR_ROW, END_HDR_ROW, case_num_col, headers)
        if hdr_row is None:
            # in DAHMO-R-wiki-20260415.xlsx sheet DAHMO R-6193-12 summary the headers are "Case #, link" in column A
            # and then "Availability" in column D
            headers = [AVAILABILITY_HDR]
            hdr_row, _ = find_header_in_col(ws, START_HDR_ROW, END_HDR_ROW, chr(ord(case_num_col) + 2), headers)
            if hdr_row is None:
                return False, case_num_col, 0

    return True, case_num_col, hdr_row


def case_like_headers_in_row(ws, import_message, header_must_appear):
    #the header row can be 5, 6 or 7
    headers = [FIRST_OPUS_HDR1, FIRST_OPUS_HDR2, FIRST_OPUS_HDR3, FIRST_OPUS_HDR4, FIRST_OPUS_HDR5]
    case_num_col = "A"
    hdr_row, header_found = find_header_in_col(ws, START_HDR_ROW, END_HDR_ROW, case_num_col, headers)
    possible_hyphen_in_case_num = False
    old_numbering_column = "A"
    case_amount_column = case_num_col
    contains_volumes = False
    additional_column = False
    if hdr_row is None:
        # in CDIAK-wiki-20260510.xlsx sheet "CDIAK 1600-1" this header cell is empty,
        # so let us try the second header
        second_num_col = "B"
        hdr_row, header_found = find_header_in_col(ws, START_HDR_ROW, END_HDR_ROW, second_num_col, [SECOND_OPUS_HDR])
    if hdr_row is None:
        case_num_col = chr(ord(case_num_col) + 1)
        case_amount_column = case_num_col
        hdr_row, header_found = find_header_in_col(ws, START_HDR_ROW, END_HDR_ROW, case_num_col, headers)
        if hdr_row is None:
            #could it be a sheet with volumes and not cases?
            contains_volumes, case_num_col, hdr_row = sheet_contains_volumes(ws)
            if contains_volumes:
                message = f"Sheet {ws.title} contains volumes"
                import_message = add_to_message(import_message, message)
                _logger.warning(message)
            elif header_must_appear:
                raise ValueError(f"No '{FIRST_OPUS_HDR1}' header found in sheet {ws.title}")
    elif header_found == FIRST_OPUS_HDR1:
        # in DAHMO-R-wiki-20260415.xlsx sheet DAHMO R-6193-12 summary the headers are "Case #, link" in column A
        # and then "Availability" in column D
        headers = [AVAILABILITY_HDR]
        availability_hdr_row, _ = find_header_in_col(ws, START_HDR_ROW, END_HDR_ROW, chr(ord(case_num_col) + 3),
                                                     headers)
        if availability_hdr_row is not None:
            contains_volumes = True

    if header_found == FIRST_OPUS_HDR4:
        possible_hyphen_in_case_num = True
        old_numbering_column = chr(ord(case_num_col) - 1)
        case_amount_column = old_numbering_column
        message = "New case numbering"
        import_message = add_to_message(import_message, message)
        _logger.warning(message)

    if header_found == FIRST_OPUS_HDR3:
        #in DAHMO-R-wiki-20260415.xlsx sheet R-6193-12 cases 5341-10422 there is an extra column "Other case #"
        cell_addr = f"{"B"}{hdr_row}"
        contents = get_cell_value(ws[cell_addr], True)
        if SECOND_VOLUME_HDR == contents:
            additional_column = True

    return (possible_hyphen_in_case_num, case_amount_column, case_num_col, hdr_row, import_message,
            old_numbering_column, contains_volumes, additional_column, header_found)


def opus_like_sheet(ws):
    (_, _, _, _, _, _, _, _, header_found) = case_like_headers_in_row(ws, '', False)
    return header_found == FIRST_OPUS_HDR1


def add_opus_page(page_table, ws, r, title, url, label, change_date, timestamp, source_type,
                  parent_name, parent_url, level, availability, import_message):
    import_message, doc_links = get_doc_links(ws, f"B{r}", import_message, url, True)
    add_page({
        "title": title,
        "url": url,
        "label": label,
        "seq_label": sequential_page_label(label),
        "level": level,
        "change_date": change_date,
        "timestamp": timestamp,
        "description": get_cell_value(ws[f"B{r}"]),
        "years": get_cell_value(ws[f"C{r}"]),
        "availability": availability,
        "doc_links": doc_links,
        "source_type": source_type,
        "parent": parent_name,
        "parent_url": parent_url,
        "comments": get_cell_value(ws[f"O{r}"]),
        "import_message": import_message,
    }, page_table)


def add_fund_page_if_necessary(ws, wiki_spreadsheet, archive_title, archive_url, title, latin_title, url,
                               source_type, r, page_table, change_date, timestamp):
    """ We only add linked pages."""
    availability = get_cell_value(ws[f"D{r}"])

    if availability != 'linked' and not url:
        # we do not add such pages
        _logger.warning(f"No link in sheet='{ws.title}', fund {title} - skipping")
        return

    comments = get_cell_value(ws[f"O{r}"])
    label = general_page_label(latin_title, title, wiki_spreadsheet)
    import_message, doc_links = get_doc_links(ws, f"B{r}", "", url, True)
    add_page({
        "title": title,
        "url": url,
        "label": label,
        "seq_label": sequential_page_label(label),
        "level": "fond",
        "change_date": change_date,
        "timestamp": timestamp,
        "description": get_cell_value(ws[f"B{r}"]),
        "years": get_cell_value(ws[f"C{r}"]),
        "availability": availability,
        "doc_links": doc_links,
        "source_type": source_type,
        "import_message": import_message,
        "parent": archive_title,
        "parent_url": archive_url,
        "comments": comments,
    }, page_table)


def process_archive_sheet(ws, wiki_spreadsheet, archive_latin_name, archive_cyrillic_name,
                          change_date_col, ref_date_col, page_table=None):
    if not page_table:
        page_table = {}
    archive_name, _, archive_url, import_message, comments = get_url_and_parent_title(ws, wiki_spreadsheet,
                                                                            archive_latin_name, archive_cyrillic_name)
    if wiki_spreadsheet:
        archive_cyrillic_name = archive_name
    source_type = get_source_type(ws)
    change_date, timestamp = get_dates(ws, archive_name, change_date_col, ref_date_col)
    label = general_page_label(archive_latin_name, archive_name, wiki_spreadsheet)
    import_message, doc_links = get_cell_link_or_str(ws, "B4", import_message)
    level = "archive"
    if opus_like_sheet(ws):
        level = "opus"
        import_message = add_to_message(import_message, f"sheet '{ws.title}' describes an opus not an archive")

    add_page({
        "title": archive_name,
        "url": archive_url,
        "label": label,
        "seq_label": sequential_page_label(label),
        "level": level,
        "description": get_cell_value(ws["A2"]),
        "availability": "linked",
        "change_date": change_date,
        "timestamp": timestamp,
        "doc_links": doc_links,
        "source_type": source_type,
        "import_message": import_message,
        "parent": "",
        "parent_url": "",
        "comments": comments,
    }, page_table)

    for r in range(7, ws.max_row + 1):
        fund_num_cell_addr = f"A{r}"
        fund_num_cell = ws[fund_num_cell_addr]
        fund_descr_cell = ws[f"B{r}"]
        if str(fund_num_cell.value).startswith("="):
            break
        fund_ids, url, in_first_cell = parse_cell_integers(fund_num_cell, fund_descr_cell, archive_latin_name)
        lower_level = log_strange_parsing_result(ws, r, fund_num_cell, fund_ids, url, wiki_spreadsheet, level)
        cell_to_use = fund_num_cell
        if not in_first_cell:
            cell_to_use = fund_descr_cell
        num_results = len(fund_ids)
        match num_results:
            case 0:
                # probably some text in the A column, like "General page content" -
                # we skip such lines
                continue
            case 1:
                # regular fund number
                title, latin_title = get_page_title_from_link(cell_to_use, wiki_spreadsheet,
                                                              archive_name, archive_latin_name)
                if "archive" == level:
                    add_fund_page_if_necessary(ws, wiki_spreadsheet, archive_name, archive_url, title,
                                           latin_title, url, source_type, r, page_table, change_date, timestamp)
                else:
                    curr_source_type, label = get_label_source_type(latin_title, source_type, title, wiki_spreadsheet)
                    language = get_language(ws, f"O{r}")
                    add_case_page(False, fund_num_cell, fund_num_cell_addr, "A", change_date, ord("G"), '',
                        curr_source_type, '', label, lower_level, '',
                        archive_name, archive_url, page_table, r, url, timestamp, title, wiki_spreadsheet, ws, language)
            case _:
                for fund_id in fund_ids:
                    title = f"{archive_cyrillic_name}/{fund_id}"
                    latin_title = f"{archive_name}/{fund_id}"
                    add_fund_page_if_necessary(ws, wiki_spreadsheet, archive_name, archive_url, title,
                                               latin_title, url, source_type, r, page_table, change_date, timestamp)

    return page_table, archive_cyrillic_name, archive_url


def process_fund_sheet(ws, wiki_spreadsheet, archive_url, archive_name, fund_id, fund_cyrillic_name,
                       change_date_col, ref_date_col, page_table=None):
    fund_name = f"{archive_name}/{fund_id}"
    if not page_table:
        page_table = {}
    parent_name, parent_latin_name, fund_url, import_message, comments = get_url_and_parent_title(ws, wiki_spreadsheet,
                                                                                        fund_name, fund_cyrillic_name)
    source_type = get_source_type(ws)
    change_date, timestamp = get_dates(ws, fund_name, change_date_col, ref_date_col)
    label = general_page_label(parent_latin_name, parent_name, wiki_spreadsheet)
    import_message, doc_links = get_cell_link_or_str(ws, "B4", import_message)

    add_page(
        {
            "title": parent_name,
            "url": fund_url,
            "label": label,
            "seq_label": sequential_page_label(label),
            "level": "fond",
            "description": get_cell_value(ws["A2"]),
            "availability": "linked",
            "change_date": change_date,
            "timestamp": timestamp,
            "doc_links": doc_links,
            "source_type": source_type,
            "import_message": import_message,
            "parent": parent_title_no_ns(parent_name, wiki_spreadsheet, fund_cyrillic_name),
            "parent_url": archive_url,
            "comments": comments,
        },
        page_table,
    )

    #the header row can be 5, 6 or 7
    hdr_row, _ = find_header_in_col(ws, START_HDR_ROW, END_HDR_ROW, "A", [FIRST_FUND_HDR])
    if hdr_row is None:
        raise ValueError(f"No '{FIRST_FUND_HDR}' header found in sheet {parent_name}")

    for r in range(hdr_row + 1, ws.max_row + 1):
        opus_num_cell_addr = f"A{r}"
        opus_num_cell = ws[opus_num_cell_addr]
        opus_descr_cell = ws[f"B{r}"]
        if str(opus_num_cell.value).startswith("="):
            break
        if opus_num_cell.value:
            curr_import_message = ""
            availability = "linked"  #default
            raw_url, in_first_cell = get_url_from_2_cells(opus_num_cell, opus_descr_cell)
            if raw_url == "":
                # no link found - cannot proceed
                _logger.warning(f"No link in sheet='{ws.title}', opus {get_cell_value(opus_num_cell)} - skipping")
                continue
            cell_to_use = opus_num_cell
            if not in_first_cell:
                cell_to_use = opus_descr_cell
                availability = "unlinked"
                _logger.warning(f"Opus {get_cell_value(opus_num_cell)} in sheet='{ws.title}', "
                                f"row={r} has no page but has a document")
                curr_import_message = "Document without a page"
            title, latin_title = get_page_title_from_link(cell_to_use, wiki_spreadsheet, fund_cyrillic_name, fund_name)
            if not title:
                _logger.warning(f"cannot determine page title: sheet='{ws.title}', row={r}  - skipping")
                continue

            necessary_parts = [fund_id, normalize_to_int_str(str(opus_num_cell.value))]
            curr_import_message = check_url_sanity(raw_url, curr_import_message, wiki_spreadsheet,
                                                   necessary_parts, ws.title, opus_num_cell_addr)
            label = general_page_label(latin_title, title, wiki_spreadsheet)
            add_opus_page(page_table, ws, r, title, raw_url, label, change_date, timestamp,
                          source_type, parent_name, fund_url, "opus", availability, curr_import_message)
    return page_table, fund_url


def process_opus_sheet(ws, wiki_spreadsheet, fund_and_opus_name, fund_and_opus_cyrillic_name, fund_name, opus_name,
                       fund_url, change_date_col, ref_date_col, page_table=None):
    if not page_table:
        page_table = {}
    opus_title, opus_latin_title, opus_url, import_message, comments = get_url_and_parent_title(ws, wiki_spreadsheet,
                                                                                      fund_and_opus_name,
                                                                                      fund_and_opus_cyrillic_name)
    label = general_page_label(opus_latin_title, opus_title, wiki_spreadsheet)
    source_type = get_source_type(ws)
    change_date, timestamp = get_dates(ws, fund_and_opus_name, change_date_col, ref_date_col)
    curr_parent_title = opus_title
    try:
        curr_parent_title = parent_title_no_ns(opus_title, wiki_spreadsheet, fund_and_opus_cyrillic_name)
    except ValueError as err:
        if "Unrecognized archive root" in str(err):
            source_type = "other"

    (possible_hyphen_in_case_num, case_amount_column, case_num_col, hdr_row, import_message,
     old_numbering_column, contains_volumes, additional_column, _) = case_like_headers_in_row(ws, import_message, True)

    #there are sometimes columns inserted between "process" and "Comments",
    #so we need to find the exact position of the latter one
    comments_col = find_header_in_row(ws, hdr_row, "G", "Z", COMMENTS_HDR)
    if comments_col is None:
        raise ValueError(f"No 'Comments' column found in sheet {opus_title}")
    description = get_cell_value(ws["A2"])
    import_message, doc_links = get_cell_link_or_str(ws, "B4", import_message)

    add_page(
        {
            "title": opus_title,
            "url": opus_url,
            "label": label,
            "seq_label": sequential_page_label(label),
            "level": "opus",
            "description": description,
            "availability": "linked",
            "change_date": change_date,
            "timestamp": timestamp,
            "doc_links": doc_links,
            "parent": curr_parent_title,
            "parent_url": fund_url,
            "source_type": source_type,
            "import_message": import_message,
            "comments": comments,
        },
        page_table,
    )

    usual_case_number_found = False
    for r in range(hdr_row + 1, ws.max_row + 1):
        case_num_cell_addr = f"{case_num_col}{r}"
        case_num_cell = ws[case_num_cell_addr]
        case_descr_cell = ws[f"{chr(ord(case_num_col) + 1)}{r}"]
        case_amount_cell = ws[f"{case_amount_column}{r}"]
        if str(case_amount_cell.value).startswith("="):
            break
        curr_import_message = ""
        ignore_parts = usual_case_number_found and not possible_hyphen_in_case_num
        if fund_and_opus_name == "DAHMO-K" and case_num_cell.value == "ELAU site":
            # there are no case numbers in DAHMO-K-wiki-20250820.xlsx
            case_number_series = [""]
            num_parts = 1
            inconsistent_parts = False
        else:
            case_number_series, num_parts, inconsistent_parts = is_series_of_case_numbers(str(case_num_cell.value),
                                                                                      ignore_parts)
        if not case_number_series:
            if possible_hyphen_in_case_num:
                curr_import_message = "Old case numbering"
                case_num_cell = ws[f"{old_numbering_column}{r}"]
                case_number_series, num_parts, inconsistent_parts = is_series_of_case_numbers(str(case_num_cell.value),
                                                                                              ignore_parts)
            if not case_number_series:
                # sometimes there is some text in the A column, like "INFORMATION AND INSTRUCTIONAL DEPARTMENT" -
                # we skip such lines
                continue
        level = "volume" if contains_volumes or num_parts == 2 else "case"
        if num_parts == 1:
            usual_case_number_found = True

        raw_url, in_first_cell = get_url_from_2_cells(case_num_cell, case_descr_cell)
        if raw_url == "":
            # no link found - cannot proceed
            continue
        if not in_first_cell and case_amount_column != case_num_col:
            case_amount_cell = ws[f"{case_amount_column}{r}"]
            if case_amount_cell.hyperlink is not None:
                #CDIAK-wiki-20260409.xlsx sheet CDIAK 1164-1 contains old and new case numbering and the case link
                # may be in both columns
                hyperlink = case_amount_cell.hyperlink
                raw_url = unquote(hyperlink.target.strip())
                in_first_cell = True
        if not in_first_cell:
            if additional_column:
                curr_import_message = "Other case numbering"
            else:
                curr_import_message = "Document without a page"

        if case_num_cell.value:
            if case_num_cell.hyperlink and case_num_cell.hyperlink.target.strip() == case_num_cell.value:
                _logger.warning(f"Apparently cells in row {r} are empty but have link {case_num_cell.value} - skipping")
                continue

            latin_title, title = get_case_title(raw_url, curr_parent_title, opus_name, case_num_cell, wiki_spreadsheet,
                                                possible_hyphen_in_case_num, fund_and_opus_name,
                                                fund_and_opus_cyrillic_name)
            if title is None:
                # If it is a comment, like "Index files linked at top of page" - skip this line
                continue

            curr_source_type, label = get_label_source_type(latin_title, source_type, title, wiki_spreadsheet)

            #title sanity check
            identical, modified_title = title_cell_val_identical(title, str(case_num_cell.value).strip(),
                                                                 num_parts, inconsistent_parts)
            if not identical:
                case_id = get_case_id(case_num_cell.value)
                new_message = f"{level} {case_id} links to {title}"
                curr_import_message = add_to_message(curr_import_message, new_message)

                if not possible_hyphen_in_case_num and level == "case":
                    title = modified_title
                _logger.warning(f"{curr_import_message}")

            if level == "volume":
                if possible_hyphen_in_case_num:
                    availability = get_cell_value(ws[f"D{r}"])
                else:
                    availability = "linked"
                if "" == curr_import_message:
                    curr_import_message = f"Volume in {title}"
                    _logger.warning(curr_import_message)
                add_opus_page(page_table, ws, r, title, raw_url, label, change_date, timestamp, source_type, opus_title,
                              fund_url, level, availability, curr_import_message)
            else:
                language = get_language(ws, f"{chr(comments_col + 8)}{r}")
                add_case_page(additional_column, case_num_cell, case_num_cell_addr, case_num_col, change_date,
                    comments_col, curr_import_message, curr_source_type, fund_name, label, level, opus_name,
                    opus_title, opus_url, page_table, r, raw_url, timestamp, title, wiki_spreadsheet, ws, language)

    return page_table


def get_label_source_type(latin_title, source_type, title, wiki_spreadsheet):
    curr_source_type = source_type
    label = general_page_label(latin_title, title, wiki_spreadsheet)
    if label is None:
        curr_source_type = "other"
        label = latin_title
    return curr_source_type, label


def add_case_page(additional_column: bool, case_num_cell, case_num_cell_addr: str, case_num_col: str,
                  change_date: str | None, comments_col: int, curr_import_message: Literal[
                                                                                       '', 'Document without a page', 'Old case numbering', 'Other case numbering'] | Any,
                  curr_source_type, fund_name, label: str | Any, level, opus_name,
                  opus_title: str | None | Any, opus_url: str, page_table: dict[Any, Any] | Any, r: int, raw_url: str,
                  timestamp: str | None, title, wiki_spreadsheet, ws, language):
    necessary_parts = [fund_name, opus_name, normalize_to_int_str(str(case_num_cell.value))]
    curr_import_message = check_url_sanity(raw_url, curr_import_message, wiki_spreadsheet,
                                           necessary_parts, ws.title, case_num_cell_addr)

    comments_cell_addr = f"{chr(comments_col)}{r}"  # default f"G{r}"
    processor_cell_addr1 = f"{chr(comments_col + 2)}{r}"  # default f"I{r}"
    processor_cell_addr2 = f"{chr(comments_col + 5)}{r}"  # default f"L{r}"
    pages_processed_cell_addr1 = f"{chr(comments_col + 3)}{r}"  # default f"J{r}"
    pages_processed_cell_addr2 = f"{chr(comments_col + 6)}{r}"  # default f"M{r}"
    when_transcribed_cell_addr = f"{chr(comments_col + 7)}{r}"  # default f"N{r}"

    description_col_offs = DESCRIPTION_COL_OFFS
    years_col_offs = YEARS_COL_OFFS
    doc_type_col_offs = DOC_TYPE_COL_OFFS
    content_code_col_offs = CONTENT_CODE_COL_OFFS
    process_code_col_offs = PROCESS_CODE_COL_OFFS
    if additional_column:
        description_col_offs = description_col_offs + 1
        years_col_offs = years_col_offs + 1
        doc_type_col_offs = doc_type_col_offs + 1
        content_code_col_offs = content_code_col_offs + 1
        process_code_col_offs = process_code_col_offs + 1

    process_code = get_cell_value(ws[f"{chr(ord(case_num_col) + process_code_col_offs)}{r}"], True)  # F
    process_code, curr_import_message = normalize_process_code(process_code, curr_import_message)
    when_transcribed = ws[when_transcribed_cell_addr].value
    when_transcribed = str(when_transcribed) if when_transcribed else ""

    curr_import_message, doc_links = get_doc_links(
        ws, f"{chr(ord(case_num_col) + description_col_offs)}{r}", curr_import_message, raw_url, False)

    add_page({
        "title": title,
        "url": raw_url,
        "label": label,
        "seq_label": sequential_page_label(label),
        "level": level,
        "change_date": change_date,
        "timestamp": timestamp,
        "description": get_cell_value(ws[f"{chr(ord(case_num_col) + description_col_offs)}{r}"]),
        "years": get_cell_value(ws[f"{chr(ord(case_num_col) + years_col_offs)}{r}"]),
        "source_type": curr_source_type,
        "parent": opus_title,
        "parent_url": opus_url,
        "doc_links": doc_links,  # B
        "doc_type": get_cell_value(ws[f"{chr(ord(case_num_col) + doc_type_col_offs)}{r}"], True),  # D
        "content_code": get_cell_value(ws[f"{chr(ord(case_num_col) + content_code_col_offs)}{r}"], True),  # E
        "process_code": process_code,
        "import_message": curr_import_message,
        # the columns for the following cells are not fixed
        "comments": get_cell_value(ws[comments_cell_addr]),
        "processor": combine_cell_value(ws[processor_cell_addr1], ws[processor_cell_addr2]),
        "pages_processed": get_cell_int_value(ws[pages_processed_cell_addr1]) +
                           get_cell_int_value(ws[pages_processed_cell_addr2]),
        "when_transcribed": when_transcribed,
        "language": language,
    }, page_table)


def normalize_process_code(process_code: str | None, import_message: str) -> tuple[str | None, str]:
    # Remove trailing tab ('\\t') from the given string if present.
    if process_code is not None:
        if process_code.endswith('\t'):
            process_code = process_code[:-1]

        valid_options = ["C1", "C2", "C3", "DP", "FX", "NO", "P1", "P2"]
        error_code = "E"
        if process_code not in valid_options:
            message = f"{process_code} is not a valid process code, changed to {error_code}"
            _logger.warning(message)
            import_message = add_to_message(import_message, message)
            process_code = error_code

    return process_code, import_message


def process_worksheets(worksheets, wiki_spreadsheet, archive_name, archive_cyrillic_name, page_table=None):
    if not page_table:
        page_table = {}
    hdr_row = 1
    archive_url = ''  #we assume the archive sheet is the first in the spreadshee
    fund_url = ''  #we assume the opus sheets come after the corresponding fund sheet
    for sheet in worksheets:
        try:
            _logger.info(f"processing worksheet: {sheet.title}")
            change_date_col = find_header_in_row(sheet, hdr_row, "I", "Z", CHANGE_DATE_HDR)
            if change_date_col is None:
                raise ValueError("No 'change date:' column, cannot proceed")
            ref_date_col = find_header_in_row(sheet, hdr_row, chr(change_date_col + 2), "Z", REF_DATE_HDR)
            if ref_date_col is None:
                if archive_name == "DAHO" or archive_name == "DAOO":
                    # Juliana's recommendation for these specific spreadsheets
                    ref_date_col = change_date_col + 3
                    cell_addr = f"{chr(ref_date_col + 1)}1"
                    date = get_cell_value(sheet[cell_addr])
                    _logger.warning(f"No 'change date:' column, proceeding anyway with date {date}")
                else:
                    raise ValueError("No 'reference date:' column")

            num_attributes, fund_opus_attributes = count_fund_opus_attributes(sheet, change_date_col)
            archive_unit_name = archive_name  #unit is either archive, or fund, or opus
            archive_unit_cyrillic_name = archive_cyrillic_name
            for attribute in fund_opus_attributes:
                archive_unit_name = f"{archive_unit_name}/{attribute}"
                archive_unit_cyrillic_name = f"{archive_unit_cyrillic_name}/{attribute}"

            # for DAHMO-K-wiki-20250820.xlsx, there are no attributes on the opus page
            header = get_cell_value(sheet['B6'])
            if 'Case description, file link' == header:
                num_attributes = 2
                fund_opus_attributes = ['', '']

            match num_attributes:
                case 2:
                    fund_name = fund_opus_attributes[0]
                    opus_name = fund_opus_attributes[-1]
                    page_table = process_opus_sheet(sheet, wiki_spreadsheet, archive_unit_name,
                                                    archive_unit_cyrillic_name, fund_name, opus_name, fund_url,
                                                    change_date_col, ref_date_col, page_table)
                case 1:
                    fund_name = fund_opus_attributes[0]
                    page_table, fund_url = process_fund_sheet(sheet, wiki_spreadsheet, archive_url, archive_name,
                                                              fund_name, archive_unit_cyrillic_name, change_date_col,
                                                              ref_date_col, page_table)
                case 0:
                    page_table, archive_cyrillic_name, archive_url = process_archive_sheet(sheet, wiki_spreadsheet,
                                                                                           archive_unit_name,
                                                                                           archive_unit_cyrillic_name,
                                                                                           change_date_col,
                                                                                           ref_date_col, page_table)
                case _:
                    raise ValueError(f"{num_attributes} for worksheet {sheet.title}")

        except Exception as e:
            e.args = (f"{e} | error in sheet {sheet.title}", *e.args[1:])
            raise  # re-raise same exception with modified message
    return page_table


def import_spreadsheet(sw_filepath, actually_write=True):
    db = Database()
    workbook = load_workbook(sw_filepath)
    wiki_spreadsheet = "-wiki" in sw_filepath.lower()
    substring = "-archive"
    if wiki_spreadsheet:
        substring = "-wiki"

    last_slash_pos = max(sw_filepath.rfind('/'), sw_filepath.rfind('\\'))
    if last_slash_pos != -1:
        file_name = sw_filepath[last_slash_pos + 1:]
    else:
        file_name = sw_filepath
    archive_substring_position = file_name.find(substring)
    if archive_substring_position != -1:
        archive_name = file_name[0: archive_substring_position]
    else:
        _logger.error(f"Spreadsheet file name {file_name} contains neither 'wiki' nor 'archive'")
        return
    if wiki_spreadsheet:
        archive_cyrillic_name = archive_name  # default for wiki
    else:
        archive_cyrillic_name = archive_name_to_cyrillic(archive_name)

    inp_page_data = process_worksheets(workbook, wiki_spreadsheet, archive_name, archive_cyrillic_name)

    if not actually_write:
        return

    if not isinstance(inp_page_data, dict):
        raise TypeError("inp_page_data is not a dictionary")
    page_data = list(inp_page_data.values())

    doc_only_fields_all_caps = {
        "doc_type",
        "content_code",
        "process_code",
    }
    doc_only_fields = doc_only_fields_all_caps | {
        "pages_processed",
        "processor",
        "when_transcribed",
        "language",
    }

    doc_and_page_fields = {
        "comments",
        "timestamp",
        "availability",
        "import_message",
        "last_imported",
    }

    # Step 1: validate all page records, remove any invalid ones
    output_pages = []
    _logger.info(f"Upserting: {len(page_data)} pages")

    for page in page_data:
        title = page.get("title")
        if not title:
            _logger.warning(f"Cannot import page with no title: parent={page.get('parent')}, label={page.get('label')}")
            continue

        # upload the page record
        page_payload = {k: v for k, v in page.items() if k not in doc_only_fields}
        #page_payload["import_message"] = "" # clear any pre-existing import messages
        abort_page = False
        while True:
            try:
                # attempt to encode page payload to verify correctness
                db.encode_records("Pages", [page_payload])
                break
            except InvalidFieldValue as err:
                _logger.error(f"malformed value: {err}, title: {page_payload["title"]}")
                if not err.value:
                    abort_page = True
                    break
                # replace unrecognized value and save a note in import_message
                page_payload[err.field] = ""
                prior_msg = page_payload.get("import_message")
                msg = f"malformed: {err.field}: {err.value}"
                page_payload["import_message"] = add_to_message(prior_msg, msg)
                # retry record normalization

        if abort_page:
            _logger.error(f"skipping malformed page record: {title}")
            continue
        output_pages.append(page_payload)

    # Step 2: batch write the page records and preserve id
    _logger.info(f"Writing {len(output_pages)} page records to db")
    output_ids = db.write("Pages", output_pages)
    for page, record_id in zip(output_pages, output_ids):
        page["Id"] = record_id

    # Step 3: form dict by page title
    page_dict = {page["url"]: page for page in output_pages}

    # Step 4: locate all parents and link to children
    parent_dict = {}
    for page in page_dict.values():
        parent_url = page.get("parent_url")
        if parent_url:
            parent_page = parent_dict.get(parent_url)
            if not parent_page:
                parent_page = page_dict.get(parent_url)
            if parent_page:
                parent_dict[parent_url] = parent_page
                child_ids = parent_page.get("child_ids", [])
                child_ids.append(page["Id"])
                parent_page["child_ids"] = child_ids

    # Step 5: link parents to their children
    for parent in parent_dict.values():
        child_ids = parent.get("child_ids")
        if child_ids:
            _logger.info(f"Linking {len(child_ids)} children for {parent['title']}")
            db.create_links(
                "Pages",
                "children",
                parent["Id"],
                child_ids)

    # Step 6: create records for all linked docs
    output_docs = {}
    doc_link_dict = {}
    for page in output_pages:
        doc_links = page.get("doc_links")
        if doc_links:
            if isinstance(doc_links, str):
                doc_links = [doc_links]
            if not isinstance(doc_links, (list, tuple)):
                raise TypeError("doc_links must be str, list or tuple")

            for doc_link in doc_links:
                # quote for document titles that contain protected characters such as "?"
                quoted_record = form_document_record(doc_link)

                doc_payload = {}
                for joint_field in doc_and_page_fields:
                    if joint_field in page:
                        doc_payload[joint_field] = page[joint_field]

                doc_payload["title"] = quoted_record["title"]
                link = quoted_record["url"]
                doc_payload["url"] = link

                relevant_page_data = inp_page_data[page["url"]]
                for k in doc_only_fields:
                    if k in relevant_page_data:
                        doc_payload[k] = relevant_page_data[k]

                for doc_field in doc_only_fields_all_caps:
                    if doc_field in doc_payload and doc_payload[doc_field] is not None:
                        doc_payload[doc_field] = doc_payload[doc_field].strip().upper()

                output_docs[link] = doc_payload
                linked_docs = doc_link_dict.get(page["url"], [])
                linked_docs.append(doc_payload)
                doc_link_dict[page["url"]] = linked_docs

    # Step 7: write doc records to db
    all_docs = list(output_docs.values())
    _logger.info(f"Writing {len(all_docs)} doc records")

    doc_ids = db.write("Documents", all_docs)
    for doc_id, doc in zip(doc_ids, all_docs):
        output_docs[doc["url"]]["Id"] = doc_id

    # Step 8: link pages to docs
    for page_url, linked_docs in doc_link_dict.items():
        page_id = page_dict[page_url]["Id"]
        page_title = page_dict[page_url]["title"]
        linked_ids = [output_docs[d["url"]]["Id"] for d in doc_link_dict[page_url]]
        _logger.info(f"Adding doc links for {page_title} ({page_id}): {linked_ids}")
        db.create_links(
            "Pages",
            "doc_links",
            page_id,
            linked_ids)

    timer.report()  # See average time spent in DB functions


def process_dir(dir_path, actually_write=True):
    dir_path = Path(dir_path)
    errors_file = dir_path / 'SummaryInfo.txt'
    with errors_file.open('a', encoding='utf-8') as f:
        for entry in dir_path.glob('*.xlsx'):
            if entry.is_file():
                start = time.perf_counter()
                msg = f"{Fore.GREEN}Processing the spreadsheet {entry}{Fore.RESET}"
                _logger.info(msg)
                f.write(msg + "\n")
                try:
                    import_spreadsheet(str(entry), actually_write)
                except Exception as e:
                    error_msg = f"Error processing {entry}: {e}\n"
                    f.write(error_msg)
                    _logger.info(f"{Fore.RED}{error_msg}{Fore.RESET}")
                end = time.perf_counter()
                f.write(f"Elapsed: {end - start:.2f} seconds\n")
                f.flush()


#testing
if __name__ == "__main__":
    filepath = "C:/jewishGen/Import2DB/SourceSpreadsheets/Wiki/ImportProblem/DAHMO-K-wiki-20250820.xlsx"
#    filepath = "C:/jewishGen/Import2DB/SourceSpreadsheets/Archives/DAHO-archive-20260213.xlsx"
    import_spreadsheet(filepath)
#    dir_path = "C:/jewishGen/Import2DB/SourceSpreadsheets/Wiki"
#    process_dir(dir_path)
